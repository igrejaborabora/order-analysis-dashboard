import os
import logging
import requests
import pandas as pd
import time
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from tenacity import retry, stop_after_attempt, wait_exponential
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import re
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor
import traceback

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('orders_analysis.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

store_name = os.getenv('SHOPIFY_STORE_NAME')
access_token = os.getenv('SHOPIFY_ACCESS_TOKEN')
storefront_token = os.getenv('SHOPIFY_STOREFRONT_TOKEN')

if not store_name or not access_token:
    messagebox.showerror("Error", "Missing SHOPIFY_STORE_NAME or SHOPIFY_ACCESS_TOKEN environment variables.")
    raise SystemExit("Missing necessary environment variables.")

if not storefront_token:
    logger.warning("SHOPIFY_STOREFRONT_TOKEN not found. International pricing might be limited.")

flores_location_id = os.getenv('LOCATION_ID_FLORES')
warehouse_location_id = os.getenv('LOCATION_ID_WAREHOUSE')

if not flores_location_id or not warehouse_location_id:
    messagebox.showerror("Error", "Missing location environment variables.")
    raise SystemExit("Missing location environment variables.")

location_ids = {
    "Pr. das Flores, 54": flores_location_id,
    "Warehouse": warehouse_location_id
}

base_url = f"https://{store_name}.myshopify.com/admin/api/2023-07/graphql.json"
rest_base_url = f"https://{store_name}.myshopify.com/admin/api/2024-01"
storefront_base_url = f"https://{store_name}.myshopify.com/api/2024-01/storefront/graphql.json"

inventory_cache: Dict[str, Dict[str, int]] = {}
cost_cache: Dict[str, float] = {}

@lru_cache(maxsize=1000)
def fetch_cost_for_inventory_item(inventory_item_id: str) -> float:
    if inventory_item_id in cost_cache:
        return cost_cache[inventory_item_id]
        
    inventory_item_id_numeric = re.search(r'\d+$', inventory_item_id).group()
    cost_url = f"{rest_base_url}/inventory_items/{inventory_item_id_numeric}.json"
    
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": access_token
    }
    
    while True:
        response = requests.get(cost_url, headers=headers)
        
        if response.status_code == 429:
            logger.warning("Rate limit exceeded. Retrying in 1 second...")
            time.sleep(1)
            continue
        
        response.raise_for_status()
        
        data = response.json()
        cost = float(data.get("inventory_item", {}).get("cost", 0.0))
        cost_cache[inventory_item_id] = cost
        return cost

@lru_cache(maxsize=1000)
def fetch_inventory_levels(inventory_item_id: str) -> Dict[str, int]:
    inventory_query = """
    query($inventoryItemId: ID!) {
      inventoryItem(id: $inventoryItemId) {
        inventoryLevels(first: 10) {
          edges {
            node {
              available
              location {
                id
                name
              }
            }
          }
        }
      }
    }
    """

    if inventory_item_id in inventory_cache:
        return inventory_cache[inventory_item_id]

    response = execute_graphql_query(inventory_query, {"inventoryItemId": inventory_item_id})
    inventory_levels = response['data']['inventoryItem']['inventoryLevels']['edges']
    
    stock_levels = {location: 0 for location in location_ids.values()}
    for level in inventory_levels:
        location_id = level['node']['location']['id']
        if location_id in stock_levels:
            stock_levels[location_id] = level['node']['available']
            
    inventory_cache[inventory_item_id] = stock_levels
    return stock_levels

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def execute_graphql_query(query: str, variables: Optional[Dict] = None) -> Dict[str, Any]:
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": access_token
    }
    
    response = requests.post(base_url, json={"query": query, "variables": variables}, headers=headers)
    response.raise_for_status()
    
    result = response.json()
    
    if 'errors' in result:
        error_message = '; '.join([error.get('message', 'Unknown error') for error in result['errors']])
        logger.error(f"GraphQL query returned errors: {error_message}")
        raise Exception(f"GraphQL errors: {error_message}")
        
    if 'data' not in result:
        logger.error("GraphQL response missing 'data' field")
        logger.error(f"Full response: {result}")
        raise Exception("Invalid GraphQL response: missing 'data' field")
        
    return result

orders_query = """
query($cursor: String) {
  orders(first: 50, after: $cursor, query: "fulfillment_status:unfulfilled AND status:open", sortKey: CREATED_AT, reverse: true) {
    edges {
      cursor
      node {
        id
        name
        createdAt
        customer {
          firstName
          lastName
        }
        shippingAddress {
          country
        }
        totalPriceSet {
          shopMoney {
            amount
            currencyCode
          }
        }
        totalShippingPriceSet {
          shopMoney {
            amount
            currencyCode
          }
        }
        displayFulfillmentStatus
        lineItems(first: 50) {
          edges {
            node {
              title
              quantity
              sku
              variant {
                id
                inventoryItem {
                  id
                }
                price
                compareAtPrice
                product {
                  vendor
                }
                weight
                weightUnit
              }
            }
          }
        }
      }
    }
    pageInfo {
      hasNextPage
      endCursor
    }
  }
}
"""

count_query = """
query {
  orders(first: 250, query: "fulfillment_status:unfulfilled AND status:open") {
    pageInfo {
      hasNextPage
      endCursor
    }
    edges {
      node {
        id
      }
    }
  }
}
"""

def fetch_unfulfilled_orders_with_progress(gui_app) -> List[Dict]:
    all_orders = []
    result = execute_graphql_query(count_query)
    if not result or 'data' not in result or 'orders' not in result['data']:
        error_msg = "Invalid response structure from GraphQL API"
        logger.error(f"{error_msg}. Response: {result}")
        raise Exception(error_msg)
    
    initial_orders = result['data']['orders']['edges']
    total_count = len(initial_orders)
    has_more = result['data']['orders']['pageInfo']['hasNextPage']
    next_cursor = result['data']['orders']['pageInfo']['endCursor']
    
    while has_more:
        count_with_cursor = f"""
        query {{
            orders(first: 250, after: "{next_cursor}", query: "fulfillment_status:unfulfilled AND status:open") {{
                pageInfo {{
                    hasNextPage
                    endCursor
                }}
                edges {{
                    node {{
                        id
                    }}
                }}
            }}
        }}
        """
        result = execute_graphql_query(count_with_cursor)
        if not result or 'data' not in result or 'orders' not in result['data']:
            break
        batch = result['data']['orders']['edges']
        total_count += len(batch)
        has_more = result['data']['orders']['pageInfo']['hasNextPage']
        next_cursor = result['data']['orders']['pageInfo']['endCursor']
    
    gui_app.total_orders = total_count
    gui_app.order_count_value.config(text=f"0/{gui_app.total_orders}")
    
    cursor = None
    while True:
        variables = {"cursor": cursor} if cursor else {}
        logger.info(f"Fetching orders with cursor: {cursor}")
        result = execute_graphql_query(orders_query, variables)
        
        if not result or 'data' not in result or 'orders' not in result['data']:
            break
        
        orders_data = result['data']['orders']
        if not orders_data.get('edges'):
            break
        
        all_orders.extend(orders_data['edges'])
        
        gui_app.processed_orders = len(all_orders)
        gui_app.order_count_value.config(text=f"{gui_app.processed_orders}/{gui_app.total_orders}")
        progress = (len(all_orders) / total_count * 100) if total_count > 0 else 0
        gui_app.fetch_progress_var.set(progress)
        gui_app.fetch_progress_label.config(text=f"{int(progress)}%")
        gui_app.root.update_idletasks()
        
        if not orders_data['pageInfo']['hasNextPage']:
            break
        cursor = orders_data['pageInfo']['endCursor']
        time.sleep(0.5)
    
    return all_orders

def check_stock_status(stock: int, required_quantity: int) -> str:
    if stock == 0:
        return "OUT OF STOCK"
    elif stock < required_quantity:
        return "NOT ENOUGH"
    else:
        return "OK"

def format_excel_output(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    
    for idx, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        header_cell = worksheet.cell(row=1, column=idx)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        max_length = max(len(str((worksheet.cell(row=r, column=idx).value) or '')) for r in range(1, len(df)+2))
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width

def apply_excel_formatting(workbook):
    # Define colors
    header_color = 'CE5870'
    colors = ['F2D5DB', 'FAEAED']  # Alternating colors for orders
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Set row height for header and format it
        worksheet.row_dimensions[1].height = 50
        
        # Format header row
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_order = None
        color_index = 0
        
        for row in range(2, worksheet.max_row + 1):
            order_number = worksheet.cell(row=row, column=1).value
            if order_number != current_order:
                current_order = order_number
                color_index = (color_index + 1) % 2
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Format 'Margin %' column as percentage if needed
                header_cell = worksheet.cell(row=1, column=col)
                if header_cell.value == 'Margin %':
                    if cell.value is not None and row > 1:
                        try:
                            value = float(str(cell.value).replace('%',''))
                            cell.value = value / 100.0
                            cell.number_format = '0.00%'
                        except (ValueError, TypeError):
                            pass
        
        # Adjust column widths
        for col in range(1, worksheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                try:
                    if cell.value is not None:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

def save_excel_file(file_path: str, df_group1: pd.DataFrame, df_group2: pd.DataFrame, df_group3: pd.DataFrame) -> None:
    try:
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        
        df_group1.to_excel(writer, sheet_name='Group 1 - Full Stock', index=False)
        df_group2.to_excel(writer, sheet_name='Group 2 - Partial Stock', index=False)
        df_group3.to_excel(writer, sheet_name='Group 3 - No Stock', index=False)
        
        workbook = writer.book
        apply_excel_formatting(workbook)
        writer.close()
        
        logger.info(f"Excel file saved successfully: {file_path}")
        messagebox.showinfo("Success", "Analysis report saved successfully!")
        os.startfile(os.path.abspath(file_path))
    except Exception as e:
        err_trace = traceback.format_exc()
        logger.error(f"Error saving Excel file: {str(e)}\n{err_trace}")
        messagebox.showerror("Error", f"An error occurred while saving the file: {str(e)}")
        raise

def determine_zone(country: str) -> str:
    country = country.strip().lower()
    if country == 'united states':
        return 'USA'
    elif country == 'france':
        return 'Europe'
    elif country in ['canada', 'australia', 'new zealand']:
        return 'RestWorld'
    else:
        return 'RestWorld'

def get_azul_cost(total_weight_kg: float, zone: str) -> float:
    adjusted_weight = total_weight_kg * 1.10
    intervals = [
        (0.019, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.049, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.099, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.249, {"Europe":5.65,"RestWorld":6.00,"USA":6.66}),
        (0.499, {"Europe":7.55,"RestWorld":10.14,"USA":11.75}),
        (0.999, {"Europe":10.65,"RestWorld":17.8,"USA":19.8}),
        (2.000, {"Europe":16.75,"RestWorld":28.00,"USA":30.35})
    ]

    for ub, costs in intervals:
        if adjusted_weight <= ub:
            return costs.get(zone, costs["RestWorld"])
    last = intervals[-1][1]
    return last.get(zone, last["RestWorld"])


def get_variant_data(session, variant_id):
    url = f"https://{store_name}.myshopify.com/admin/api/2023-01/variants/{variant_id}.json"
    response = session.get(url)
    if response.status_code == 200:
        return response.json()
    return None

def process_orders(orders: List[Dict]) -> List[Dict]:
    processed_data = []
    batch_size = 50
    session = requests.Session()
    session.headers.update({
        'X-Shopify-Access-Token': access_token,
        'Content-Type': 'application/json'
    })

    for batch_start in range(0, len(orders), batch_size):
        batch_orders = orders[batch_start:batch_start+batch_size]
        
        inventory_items = set()
        for order_edge in batch_orders:
            order = order_edge['node']
            for item_edge in order['lineItems']['edges']:
                item = item_edge['node']
                inventory_items.add(item['variant']['inventoryItem']['id'])
        
        with ThreadPoolExecutor(max_workers=10) as executor:
            list(executor.map(fetch_cost_for_inventory_item, inventory_items))
            list(executor.map(fetch_inventory_levels, inventory_items))
        
        for order_edge in batch_orders:
            order = order_edge['node']
            logger.info(f"Processing Order: {order['name']}")

            country = order['shippingAddress']['country'] if order.get('shippingAddress') else ''
            zone = determine_zone(country)
            order_currency = order['totalPriceSet']['shopMoney']['currencyCode']
            total_value = float(order['totalPriceSet']['shopMoney']['amount'] or 0.0)

            charged_transport_cost = 0.0
            if order.get('totalShippingPriceSet') and order['totalShippingPriceSet'].get('shopMoney'):
                shipping_amount = order['totalShippingPriceSet']['shopMoney'].get('amount')
                if shipping_amount:
                    charged_transport_cost = float(shipping_amount)
            
            total_weight = 0.0
            sum_product_profit = 0.0
            
            line_items_data = []
            for item_edge in order['lineItems']['edges']:
                item = item_edge['node']
                inventory_item_id = item['variant']['inventoryItem']['id']
                cost_price = fetch_cost_for_inventory_item(inventory_item_id)
                quantity = item['quantity']
                final_price = float(item['variant']['price'] or 0.0)
                price_diff = final_price - cost_price

                variant_weight = float(item['variant'].get('weight', 0.0))
                total_weight += variant_weight * quantity
                sum_product_profit += (price_diff * quantity)
                
                line_items_data.append((item, cost_price, final_price, price_diff, quantity))
            
            real_transport_cost = get_azul_cost(total_weight, zone)
            
            transport_cost_difference = real_transport_cost - charged_transport_cost
            adjusted_total = total_value - charged_transport_cost if total_value != 0 else 0
            if adjusted_total != 0:
                margin_percent = 1 - ((sum_product_profit + transport_cost_difference) / adjusted_total)
            else:
                margin_percent = 0.0
            
            for (item, cost_price, final_price, price_diff, quantity) in line_items_data:
                inventory_item_id = item['variant']['inventoryItem']['id']
                stock_levels = fetch_inventory_levels(inventory_item_id)
                stock_flores = stock_levels[location_ids["Pr. das Flores, 54"]]
                stock_status = check_stock_status(stock_flores, quantity)

                variant_id = item['variant']['id']
                ean = ''
                if variant_id:
                    numeric_id = variant_id.split('/')[-1]
                    variant_data = get_variant_data(session, numeric_id)
                    if variant_data and 'variant' in variant_data:
                        ean = variant_data['variant'].get('barcode', '')

                processed_data.append({
                    "Order Number": order['name'],
                    "Order Date": order['createdAt'],
                    "Country": country,
                    "Currency Code": order_currency,
                    "SKU": item['sku'],
                    "EAN": ean,
                    "Product": item['title'],
                    "Quantity": quantity,
                    "Vendor": item['variant']['product']['vendor'],
                    "Cost_Per_Item": round(cost_price, 2),
                    "Unit_Price": round(final_price, 2),
                    "Compare_Price": round(float(item['variant']['compareAtPrice'] or 0.0), 2),
                    "Price_Diff": round(price_diff, 2),
                    "Real_Profit": round(price_diff * quantity, 2),
                    "Margin_Percent": round(margin_percent, 4),
                    "Total Value": round(total_value, 2),
                    "Stock_Status": stock_status,
                    "Stock_Flores": stock_flores,
                    "Stock_Warehouse": stock_levels[location_ids["Warehouse"]],
                    "Numeric_Weight": variant_weight,
                    "Transport_Cost": round(charged_transport_cost, 2),
                    "Real_Transport_Cost": round(real_transport_cost, 2),
                    "Zone": zone,
                    "group": ""
                })
                
    return processed_data

def convert_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    def margin_to_percent(x):
        if isinstance(x, str) and '%' in x:
            return x
        try:
            x_dec = str(x).replace(',', '.')
            val = float(x_dec) * 100.0
            return f"{val:.2f}%".replace('.', ',')
        except (ValueError, TypeError):
            return x

    def value_to_decimal(x):
        try:
            if isinstance(x, str):
                return float(x.replace(',', '.'))
            return float(x)
        except (ValueError, TypeError):
            return x

    numeric_columns = [
        "Total Value", "Transport_Cost", "Cost Per Item", "Unit Price",
        "Compare Price", "Price Diff", "Real Profit", "Margin %"
    ]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].apply(value_to_decimal)
            df[col] = df[col].apply(lambda x: f"{x:.2f}".replace('.', ',') if isinstance(x, (int, float)) else x)

    if "Margin %" in df.columns:
        df["Margin %"] = df["Margin %"].apply(margin_to_percent)

    return df

def add_total_weight_and_real_cost(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df["Total_Weight"] = ""
        return df

    df = df.copy()
    
    def safe_numeric_convert(series):
        if series.dtype == object:
            return pd.to_numeric(series.str.replace(',', '.'), errors='coerce').fillna(0)
        return pd.to_numeric(series, errors='coerce').fillna(0)
    
    df['Numeric_Weight'] = safe_numeric_convert(df['Numeric_Weight'])
    
    order_weights = df.groupby("Order Number")['Numeric_Weight'].transform('sum')
    df["Total_Weight"] = order_weights
    df["Total_Weight"] = df["Total_Weight"].apply(lambda x: f"{x:.3f}".replace('.', ','))
    
    return df

def split_data_by_group(processed_data: List[Dict]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = pd.DataFrame(processed_data)
    
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    df['Stock_Status'] = df['Stock_Status'].astype(str)
    
    df = df.copy()
    
    def determine_group(group_df):
        statuses = group_df['Stock_Status'].astype(str)
        if 'OK' in statuses.values:
            if 'OUT OF STOCK' not in statuses.values and 'NOT ENOUGH' not in statuses.values:
                return pd.Series(1, index=group_df.index)
            else:
                return pd.Series(2, index=group_df.index)
        else:
            return pd.Series(3, index=group_df.index)
    
    groups = []
    for _, group_df in df.groupby('Order Number'):
        groups.append(determine_group(group_df))
    df['group'] = pd.concat(groups)
    
    df = add_total_weight_and_real_cost(df)
    
    column_names = {
        "Order Number": "Order Number",
        "Order Date": "Order Date",
        "Country": "Country",
        "Currency Code": "Currency Code",
        "SKU": "SKU",
        "EAN": "EAN",
        "Product": "Product",
        "Total Value": "Total Value",
        "Quantity": "Quantity",
        "Vendor": "Vendor",
        "Stock_Status": "Stock Status",
        "Stock_Flores": "Stock Flores",
        "Stock_Warehouse": "Stock Warehouse",
        "Cost_Per_Item": "Cost Per Item",
        "Unit_Price": "Unit Price",
        "Compare_Price": "Compare Price",
        "Price_Diff": "Price Diff",
        "Real_Profit": "Real Profit",
        "Margin_Percent": "Margin %",
        "Transport_Cost": "Transport Cost",
        "Total_Weight": "Total Weight",
        "Real_Transport_Cost": "Real Transport Cost",
        "group": "Group"
    }
    
    df = df.rename(columns=column_names)
    
    column_order = [
        "Order Number", "Order Date", "Country", "Currency Code", "SKU", "EAN", "Product",
        "Quantity", "Vendor", "Cost Per Item", "Unit Price", "Compare Price", "Price Diff",
        "Real Profit", "Margin %", "Total Value", "Stock Status", "Stock Flores",
        "Stock Warehouse", "Total Weight", "Transport Cost", "Real Transport Cost", "Group"
    ]
    
    for col in column_order:
        if col not in df.columns:
            print(f"Missing column: {col}")
    
    df = df[column_order]
    
    df_group1 = df[df['Group'] == 1].copy()
    df_group2 = df[df['Group'] == 2].copy()
    df_group3 = df[df['Group'] == 3].copy()
    
    df_group1 = convert_numeric_columns(df_group1)
    df_group2 = convert_numeric_columns(df_group2)
    df_group3 = convert_numeric_columns(df_group3)
    
    return df_group1, df_group2, df_group3

def setup_treeview_style():
    style = ttk.Style()
    style.theme_use("clam")

    style.configure("MyTreeview.Heading",
                    background="#CCCCCC",
                    foreground="black",
                    font=('Arial', 10, 'bold'))

    style.configure("MyTreeview",
                    background="white",
                    foreground="black",
                    font=('Arial', 10),
                    rowheight=25,
                    fieldbackground="white")

    style.layout("MyTreeview", style.layout("Treeview"))
    return style

def configure_columns(tree: ttk.Treeview):
    columns = [
        "Order Number","Order Date","Country","Currency Code","SKU","EAN","Product","Total Value","Quantity","Vendor",
        "Stock_Status","Stock_Flores","Stock_Warehouse","Cost_Per_Item","Unit_Price","Compare_Price","Price_Diff",
        "Margin_Percent","Transport_Cost","Total_Weight","Real_Transport_Cost","Real_Profit","group"
    ]
    tree["columns"] = columns
    tree["show"] = "headings"
    
    col_config = {
        "Order Number": (110, 'center'),
        "Order Date": (180, 'center'),
        "Country": (100, 'center'),
        "Currency Code": (100, 'center'),
        "SKU": (120, 'center'),
        "EAN": (100, 'center'),
        "Product": (300, 'w'),
        "Total Value": (80, 'center'),
        "Quantity": (70, 'center'),
        "Vendor": (100, 'center'),
        "Stock_Status": (100, 'center'),
        "Stock_Flores": (100, 'center'),
        "Stock_Warehouse": (120, 'center'),
        "Cost_Per_Item": (90, 'center'),
        "Unit_Price": (80, 'center'),
        "Compare_Price": (90, 'center'),
        "Price_Diff": (90, 'center'),
        "Margin_Percent": (90, 'center'),
        "Transport_Cost": (90, 'center'),
        "Total_Weight": (90, 'center'),
        "Real_Transport_Cost": (90, 'center'),
        "Real_Profit": (90, 'center'),
        "group": (50, 'center')
    }

    for col in columns:
        width, align = col_config[col]
        anchor = tk.W if align == 'w' else tk.CENTER
        tree.heading(col, text=col, anchor=tk.CENTER)
        tree.column(col, width=width, anchor=anchor)

class CTTCostsViewer:
    # ... original CTT costs viewer code ...
    # Not focusing on changes here since layout request was about main GUI
    def __init__(self, parent):
        pass

class OrderAnalysisGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Order Analysis")
        self.root.geometry("1300x900")
        
        self.order_data = []
        self.total_orders = 0
        self.processed_orders = 0
        
        self.start_time = 0
        self.is_running = False
        self.processed_data = []
        self.group_data = None

        self.style = setup_treeview_style()

        self.main_frame = ttk.Frame(self.root, padding=20)
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.create_timer_and_count_widgets()
        self.create_table()
        self.create_bottom_section()
        self.create_button_frame()

        self.main_frame.columnconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def create_timer_and_count_widgets(self):
        self.timer_frame = ttk.Frame(self.main_frame, padding=10)
        self.timer_frame.grid(row=0, column=0, sticky=(tk.W, tk.E))

        self.timer_label = ttk.Label(self.timer_frame, text="Time Elapsed:", font=('Arial', 10, 'bold'))
        self.timer_label.grid(row=0, column=0, padx=30)
        
        self.timer_value = ttk.Label(self.timer_frame, text="00:00", font=('Arial', 12))
        self.timer_value.grid(row=0, column=1, padx=30)
        
        self.order_count_label = ttk.Label(self.timer_frame, text="Orders to Process:", font=('Arial', 10, 'bold'))
        self.order_count_label.grid(row=0, column=2, padx=30)
        
        self.order_count_value = ttk.Label(self.timer_frame, text="0/0", font=('Arial', 12))
        self.order_count_value.grid(row=0, column=3, padx=30)

    def create_table(self):
        self.table_frame = ttk.Frame(self.main_frame, padding=10)
        self.table_frame.grid(row=1, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.tree = ttk.Treeview(self.table_frame, style='MyTreeview', height=16)
        configure_columns(self.tree)
        
        scrollbar = ttk.Scrollbar(self.table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        self.table_frame.columnconfigure(0, weight=1)

    def create_bottom_section(self):
        # Frame that holds progress and stats side by side
        self.bottom_section = ttk.Frame(self.main_frame, padding=10)
        self.bottom_section.grid(row=2, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Progress frame on the left
        self.progress_frame = ttk.LabelFrame(self.bottom_section, text="Progress", padding=10)
        self.progress_frame.grid(row=0, column=0, sticky=(tk.W, tk.N, tk.S), padx=10, pady=10)
        
        ttk.Label(self.progress_frame, text="Fetching Orders:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.fetch_progress_var = tk.DoubleVar()
        self.fetch_progress = ttk.Progressbar(
            self.progress_frame, 
            variable=self.fetch_progress_var,
            maximum=100,
            length=150,  # half of original length
            mode='determinate'
        )
        self.fetch_progress.grid(row=0, column=1, padx=5, pady=5)
        self.fetch_progress_label = ttk.Label(self.progress_frame, text="0%")
        self.fetch_progress_label.grid(row=0, column=2, padx=5, pady=5)
        
        ttk.Label(self.progress_frame, text="Processing Data:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.compute_progress_var = tk.DoubleVar()
        self.compute_progress = ttk.Progressbar(
            self.progress_frame,
            variable=self.compute_progress_var,
            maximum=100,
            length=150, # half of original
            mode='determinate'
        )
        self.compute_progress.grid(row=1, column=1, padx=5, pady=5)
        self.compute_progress_label = ttk.Label(self.progress_frame, text="0%")
        self.compute_progress_label.grid(row=1, column=2, padx=5, pady=5)

        # Stats board on the right
        self.stats_frame = ttk.LabelFrame(self.bottom_section, text="Stats Board", padding=10)
        self.stats_frame.grid(row=0, column=1, sticky=(tk.E, tk.N, tk.S), padx=10, pady=10)

        self.stats_labels = {
            'full_orders': ttk.Label(self.stats_frame, text="Full Stock: 0 | €0.00"),
            'partial_orders': ttk.Label(self.stats_frame, text="Partial Stock: 0 | €0.00"),
            'no_stock_orders': ttk.Label(self.stats_frame, text="No Stock: 0 | €0.00"),
            'top_countries': ttk.Label(self.stats_frame, text="Top 5 Countries:\n"),
            'lateness': ttk.Label(self.stats_frame, text="Lateness:\n>3 days: 0\n>=4 days:0\n>=7 days:0\n>=9 days:0")
        }

        row_counter = 0
        self.stats_labels['full_orders'].grid(row=row_counter, column=0, sticky=tk.W, padx=5, pady=5)
        row_counter += 1
        self.stats_labels['partial_orders'].grid(row=row_counter, column=0, sticky=tk.W, padx=5, pady=5)
        row_counter += 1
        self.stats_labels['no_stock_orders'].grid(row=row_counter, column=0, sticky=tk.W, padx=5, pady=5)
        row_counter += 1
        self.stats_labels['top_countries'].grid(row=row_counter, column=0, sticky=tk.W, padx=5, pady=5)
        row_counter += 1
        self.stats_labels['lateness'].grid(row=row_counter, column=0, sticky=tk.W, padx=5, pady=5)

        self.bottom_section.columnconfigure(0, weight=1)
        self.bottom_section.columnconfigure(1, weight=1)

    def create_button_frame(self):
        self.button_frame = ttk.Frame(self.main_frame, padding=10)
        self.button_frame.grid(row=3, column=0, sticky=tk.EW)

        self.ctt_costs_button = ttk.Button(self.button_frame, text="CTT Costs", command=self.ctt_costs_action)
        self.export_button = ttk.Button(self.button_frame, text="Export Excel", command=self.export_to_excel_main)
        self.start_button = ttk.Button(self.button_frame, text="Start Analysis", command=self.start_analysis)

        self.ctt_costs_button.pack(side='left', padx=10)
        self.export_button.pack(side='right', padx=10)
        self.start_button.pack(side='right', padx=10)

    def update_timer(self):
        if self.is_running:
            elapsed_time = int(time.time() - self.start_time)
            minutes = elapsed_time // 60
            seconds = elapsed_time % 60
            self.timer_value.config(text=f"{minutes:02d}:{seconds:02d}")
            self.root.after(1000, self.update_timer)

    def update_table(self, order_data: List[Dict]):
        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            for i, row in enumerate(order_data):
                values = []
                for col in self.tree["columns"]:
                    val = row.get(col, "")
                    values.append("" if val is None else str(val))
            
                tag = "oddrow" if i % 2 else "evenrow"
                self.tree.insert("", "end", values=values, tags=(tag,))
        
            self.tree.tag_configure("oddrow", background="#F2D5DB")
            self.tree.tag_configure("evenrow", background="#FAEAED")
        
        except Exception as e:
            logger.error(f"Error in update_table: {str(e)}")
            traceback.print_exc()

    def update_stats_board(self, df_group1: pd.DataFrame, df_group2: pd.DataFrame, df_group3: pd.DataFrame):
        def get_orders_count_and_sum(df):
            if df.empty:
                return 0, 0.0
            unique_orders = df['Order Number'].unique()
            count_orders = len(unique_orders)
            total_values = df.groupby('Order Number')['Total Value'].first().sum()
            return count_orders, total_values

        full_count, full_sum = get_orders_count_and_sum(df_group1)
        partial_count, partial_sum = get_orders_count_and_sum(df_group2)
        nostock_count, nostock_sum = get_orders_count_and_sum(df_group3)

        self.stats_labels['full_orders'].config(text=f"Full Stock: {full_count} | €{full_sum:.2f}")
        self.stats_labels['partial_orders'].config(text=f"Partial Stock: {partial_count} | €{partial_sum:.2f}")
        self.stats_labels['no_stock_orders'].config(text=f"No Stock: {nostock_count} | €{nostock_sum:.2f}")

        if self.processed_data:
            df_all = pd.DataFrame(self.processed_data)
            order_country = df_all.groupby('Order Number')['Country'].first().value_counts()
            top_countries = order_country.head(5)

            top_text = "Top 5 Countries:\n"
            for i, (country, cnt) in enumerate(top_countries.items(), start=1):
                top_text += f"{i}. {country} - {cnt}\n"
            self.stats_labels['top_countries'].config(text=top_text)

            now = datetime.utcnow()
            df_orders = df_all.groupby('Order Number')['Order Date'].first().reset_index()
            df_orders['Order Date'] = pd.to_datetime(df_orders['Order Date'])
            df_orders['days_old'] = (now - df_orders['Order Date']).dt.days

            more_3 = (df_orders['days_old'] > 3).sum()
            more_4 = (df_orders['days_old'] >= 4).sum()
            more_7 = (df_orders['days_old'] >= 7).sum()
            more_9 = (df_orders['days_old'] >= 9).sum()

            lateness_text = (f"Lateness:\n"
                             f">3 days: {more_3}\n"
                             f">=4 days: {more_4}\n"
                             f">=7 days: {more_7}\n"
                             f">=9 days: {more_9}")
            self.stats_labels['lateness'].config(text=lateness_text)
        else:
            self.stats_labels['top_countries'].config(text="Top 5 Countries:\nNo data")
            self.stats_labels['lateness'].config(text="Lateness:\nNo data")

    def run_analysis(self):
        try:
            orders = fetch_unfulfilled_orders_with_progress(self)
            logger.info("Processing orders now...")
            self.processed_data = self.process_orders_with_progress(orders)
            
            df_group1, df_group2, df_group3 = split_data_by_group(self.processed_data)
            self.group_data = {1: df_group1, 2: df_group2, 3: df_group3}

            self.update_stats_board(df_group1, df_group2, df_group3)
            
        except Exception as e:
            err_trace = traceback.format_exc()
            logger.error(f"Error during analysis: {str(e)}\n{err_trace}")
            messagebox.showerror("Error", f"An error occurred during analysis:\n{err_trace}")
        finally:
            self.is_running = False
            self.start_button.config(state='normal')

    def process_orders_with_progress(self, orders: List[Dict]) -> List[Dict]:
        processed_data = []
        batch_size = 10
        total = len(orders)
        
        for batch_idx in range(0, total, batch_size):
            batch_orders = orders[batch_idx:batch_idx+batch_size]
            try:
                batch_processed = process_orders(batch_orders)
            except Exception as e:
                err_trace = traceback.format_exc()
                logger.error(f"Error in process_orders_with_progress: {str(e)}\n{err_trace}")
                messagebox.showerror("Error", f"An error occurred in process_orders_with_progress:\n{err_trace}")
                break
            processed_data.extend(batch_processed)
            
            progress = ((batch_idx + len(batch_orders)) / total) * 100 if total else 0
            self.compute_progress_var.set(progress)
            self.compute_progress_label.config(text=f"{int(progress)}%")
            self.root.update_idletasks()
            
            self.update_table(processed_data[-16:])
        
        return processed_data

    def start_analysis(self):
        self.start_button.config(state='disabled')
        self.start_time = time.time()
        self.is_running = True
        self.update_timer()
        Thread(target=self.run_analysis, daemon=True).start()

    def export_to_excel_main(self):
        if not self.processed_data:
            messagebox.showerror("Error", "No data available to export. Please run the analysis first.")
            return
        
        try:
            df_group1, df_group2, df_group3 = split_data_by_group(self.processed_data)
            timestamp_str = datetime.now().strftime('%Y_%m_%d_%M_%S')
            file_name = f"Cosmetyque_Orders_Grouped_{timestamp_str}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Analysis Report",
                initialfile=file_name
            )
            
            if file_path:
                save_excel_file(file_path, df_group1, df_group2, df_group3)
        except Exception as e:
            err_trace = traceback.format_exc()
            logger.error(f"Error in export_to_excel_main: {str(e)}\n{err_trace}")
            messagebox.showerror("Error", f"An error occurred in export_to_excel_main:\n{err_trace}")

    def ctt_costs_action(self):
        CTTCostsViewer(self.root)

def main():
    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(('localhost', 12345))
        root = tk.Tk()
        app = OrderAnalysisGUI(root)
        root.mainloop()
    except socket.error:
        messagebox.showwarning("Warning", "Application is already running!")
    finally:
        sock.close()

if __name__ == "__main__":
    main()
