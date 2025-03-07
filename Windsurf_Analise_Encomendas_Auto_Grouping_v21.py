import os
import logging
import requests
import pandas as pd
import time
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from tenacity import retry, stop_after_attempt, wait_exponential
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import re
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor
import traceback
import socket

# Setup logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('orders_analysis.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

store_name = os.getenv('SHOPIFY_STORE_NAME')
access_token = os.getenv('SHOPIFY_ACCESS_TOKEN')
storefront_token = os.getenv('SHOPIFY_STOREFRONT_TOKEN')

if not store_name or not access_token:
    messagebox.showerror("Error", "Missing SHOPIFY_STORE_NAME or SHOPIFY_ACCESS_TOKEN environment variables.")
    raise SystemExit("Missing necessary environment variables.")

if not storefront_token:
    logger.warning("SHOPIFY_STOREFRONT_TOKEN not found. International pricing might be limited.")

flores_location_id = os.getenv('LOCATION_ID_FLORES')
warehouse_location_id = os.getenv('LOCATION_ID_WAREHOUSE')

if not flores_location_id or not warehouse_location_id:
    messagebox.showerror("Error", "Missing location environment variables.")
    raise SystemExit("Missing location environment variables.")

location_ids = {
    "Pr. das Flores, 54": flores_location_id,
    "Warehouse": warehouse_location_id
}

base_url = f"https://{store_name}.myshopify.com/admin/api/2023-07/graphql.json"
rest_base_url = f"https://{store_name}.myshopify.com/admin/api/2024-01"
storefront_base_url = f"https://{store_name}.myshopify.com/api/2024-01/storefront/graphql.json"

inventory_cache: Dict[str, Dict[str, int]] = {}
cost_cache: Dict[str, float] = {}

@lru_cache(maxsize=1000)
def fetch_cost_for_inventory_item(inventory_item_id: str) -> float:
    if inventory_item_id in cost_cache:
        return cost_cache[inventory_item_id]
        
    inventory_item_id_numeric = re.search(r'\d+$', inventory_item_id).group()
    cost_url = f"{rest_base_url}/inventory_items/{inventory_item_id_numeric}.json"
    
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": access_token
    }
    
    while True:
        response = requests.get(cost_url, headers=headers)
        
        if response.status_code == 429:
            logger.warning("Rate limit exceeded. Retrying in 1 second...")
            time.sleep(1)
            continue
        
        response.raise_for_status()
        
        data = response.json()
        cost = float(data.get("inventory_item", {}).get("cost", 0.0))
        cost_cache[inventory_item_id] = cost
        return cost

@lru_cache(maxsize=1000)
def fetch_inventory_levels(inventory_item_id: str) -> Dict[str, int]:
    inventory_query = """
    query($inventoryItemId: ID!) {
      inventoryItem(id: $inventoryItemId) {
        inventoryLevels(first: 10) {
          edges {
            node {
              available
              location {
                id
                name
              }
            }
          }
        }
      }
    }
    """

    if inventory_item_id in inventory_cache:
        return inventory_cache[inventory_item_id]

    response = execute_graphql_query(inventory_query, {"inventoryItemId": inventory_item_id})
    inventory_levels = response['data']['inventoryItem']['inventoryLevels']['edges']
    
    stock_levels = {location: 0 for location in location_ids.values()}
    for level in inventory_levels:
        location_id = level['node']['location']['id']
        if location_id in stock_levels:
            stock_levels[location_id] = level['node']['available']
            
    inventory_cache[inventory_item_id] = stock_levels
    return stock_levels

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def execute_graphql_query(query: str, variables: Optional[Dict] = None) -> Dict[str, Any]:
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": access_token
    }
    
    response = requests.post(base_url, json={"query": query, "variables": variables}, headers=headers)
    response.raise_for_status()
    
    result = response.json()
    
    if 'errors' in result:
        error_message = '; '.join([error.get('message', 'Unknown error') for error in result['errors']])
        logger.error(f"GraphQL query returned errors: {error_message}")
        raise Exception(f"GraphQL errors: {error_message}")
        
    if 'data' not in result:
        logger.error("GraphQL response missing 'data' field")
        logger.error(f"Full response: {result}")
        raise Exception("Invalid GraphQL response: missing 'data' field")
        
    return result

orders_query = """
query($cursor: String) {
  orders(first: 50, after: $cursor, query: "fulfillment_status:unfulfilled AND status:open", sortKey: CREATED_AT, reverse: true) {
    edges {
      cursor
      node {
        id
        name
        createdAt
        customer {
          firstName
          lastName
        }
        shippingAddress {
          country
        }
        totalPriceSet {
          shopMoney {
            amount
            currencyCode
          }
        }
        totalShippingPriceSet {
          shopMoney {
            amount
            currencyCode
          }
        }
        displayFulfillmentStatus
        lineItems(first: 50) {
          edges {
            node {
              title
              quantity
              sku
              variant {
                id
                inventoryItem {
                  id
                }
                price
                compareAtPrice
                product {
                  vendor
                }
                weight
                weightUnit
              }
            }
          }
        }
      }
    }
    pageInfo {
      hasNextPage
      endCursor
    }
  }
}
"""

count_query = """
query {
  orders(first: 250, query: "fulfillment_status:unfulfilled AND status:open") {
    pageInfo {
      hasNextPage
      endCursor
    }
    edges {
      node {
        id
      }
    }
  }
}
"""

def fetch_unfulfilled_orders_with_progress(gui_app) -> List[Dict]:
    all_orders = []
    result = execute_graphql_query(count_query)
    if not result or 'data' not in result or 'orders' not in result['data']:
        error_msg = "Invalid response structure from GraphQL API"
        logger.error(f"{error_msg}. Response: {result}")
        raise Exception(error_msg)
    
    initial_orders = result['data']['orders']['edges']
    total_count = len(initial_orders)
    has_more = result['data']['orders']['pageInfo']['hasNextPage']
    next_cursor = result['data']['orders']['pageInfo']['endCursor']
    
    while has_more:
        count_with_cursor = f"""
        query {{
            orders(first: 250, after: "{next_cursor}", query: "fulfillment_status:unfulfilled AND status:open") {{
                pageInfo {{
                    hasNextPage
                    endCursor
                }}
                edges {{
                    node {{
                        id
                    }}
                }}
            }}
        }}
        """
        result = execute_graphql_query(count_with_cursor)
        if not result or 'data' not in result or 'orders' not in result['data']:
            break
        batch = result['data']['orders']['edges']
        total_count += len(batch)
        has_more = result['data']['orders']['pageInfo']['hasNextPage']
        next_cursor = result['data']['orders']['pageInfo']['endCursor']
    
    gui_app.total_orders = total_count
    gui_app.order_count_value.config(text=f"0/{gui_app.total_orders}")
    
    cursor = None
    while True:
        variables = {"cursor": cursor} if cursor else {}
        logger.info(f"Fetching orders with cursor: {cursor}")
        result = execute_graphql_query(orders_query, variables)
        
        if not result or 'data' not in result or 'orders' not in result['data']:
            break
        
        orders_data = result['data']['orders']
        if not orders_data.get('edges'):
            break
        
        all_orders.extend(orders_data['edges'])
        
        gui_app.processed_orders = len(all_orders)
        gui_app.order_count_value.config(text=f"{gui_app.processed_orders}/{gui_app.total_orders}")
        progress = (len(all_orders) / total_count * 100) if total_count > 0 else 0
        gui_app.fetch_progress_var.set(progress)
        gui_app.fetch_progress_label.config(text=f"{int(progress)}%")
        gui_app.root.update_idletasks()
        
        if not orders_data['pageInfo']['hasNextPage']:
            break
        cursor = orders_data['pageInfo']['endCursor']
        time.sleep(0.5)
    
    return all_orders

def check_stock_status(stock: int, required_quantity: int) -> str:
    if stock == 0:
        return "OUT OF STOCK"
    elif stock < required_quantity:
        return "NOT ENOUGH"
    else:
        return "OK"

def format_excel_output(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    
    for idx, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        header_cell = worksheet.cell(row=1, column=idx)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        max_length = max(len(str((worksheet.cell(row=r, column=idx).value) or '')) for r in range(1, len(df)+2))
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width

def apply_excel_formatting(workbook):
    # Define colors
    header_color = 'CE5870'
    colors = ['F2D5DB', 'FAEAED']  # Alternating colors for orders
    
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        # Set row height for header and format it
        worksheet.row_dimensions[1].height = 50
        
        # Format header row
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Get unique orders and their row ranges
        current_order = None
        color_index = 0
        
        # Format data rows
        for row in range(2, worksheet.max_row + 1):
            order_number = worksheet.cell(row=row, column=1).value
            
            if order_number != current_order:
                current_order = order_number
                color_index = (color_index + 1) % 2
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=colors[color_index],
                                      end_color=colors[color_index],
                                      fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Adjust column widths
        for col in range(1, worksheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for row in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=row, column=col)
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

def save_excel_file(file_path: str, df_group1: pd.DataFrame, df_group2: pd.DataFrame, df_group3: pd.DataFrame) -> None:
    try:
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        
        df_group1.to_excel(writer, sheet_name='Group 1 - Full Stock', index=False)
        df_group2.to_excel(writer, sheet_name='Group 2 - Partial Stock', index=False)
        df_group3.to_excel(writer, sheet_name='Group 3 - No Stock', index=False)
        
        workbook = writer.book
        apply_excel_formatting(workbook)
        
        writer.close()
        
        logger.info(f"Excel file saved successfully: {file_path}")
        messagebox.showinfo("Success", "Analysis report saved successfully!")
        os.startfile(os.path.abspath(file_path))
    except Exception as e:
        err_trace = traceback.format_exc()
        logger.error(f"Error saving Excel file: {str(e)}\n{err_trace}")
        messagebox.showerror("Error", f"An error occurred while saving the file: {str(e)}")
        raise

def determine_zone(country: str) -> str:
    country = country.strip().lower()
    if country == 'united states':
        return 'USA'
    elif country == 'france':
        return 'Europe'
    elif country in ['canada', 'australia', 'new zealand']:
        return 'RestWorld'
    else:
        return 'RestWorld'

def get_azul_cost(total_weight_kg: float, zone: str) -> float:
    adjusted_weight = total_weight_kg * 1.10
    intervals = [
        (0.019, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.049, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.099, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.249, {"Europe":5.65,"RestWorld":6.00,"USA":6.66}),
        (0.499, {"Europe":7.55,"RestWorld":10.14,"USA":11.75}),
        (0.999, {"Europe":10.65,"RestWorld":17.8,"USA":19.8}),
        (2.000, {"Europe":16.75,"RestWorld":28.00,"USA":30.35})
    ]

    for ub, costs in intervals:
        if adjusted_weight <= ub:
            return costs.get(zone, costs["RestWorld"])
    last = intervals[-1][1]
    return last.get(zone, last["RestWorld"])

def get_variant_data(session, variant_id):
    print(f"Fetching variant data for ID: {variant_id}")
    url = f"https://{store_name}.myshopify.com/admin/api/2023-01/variants/{variant_id}.json"
    response = session.get(url)
    print(f"Response status: {response.status_code}")
    if response.status_code == 200:
        data = response.json()
        print(f"Variant data: {data}")
        return data
    print(f"Error response: {response.text}")
    return None

def process_orders(orders: List[Dict]) -> List[Dict]:
    processed_data = []
    batch_size = 50
    session = requests.Session()
    session.headers.update({
        'X-Shopify-Access-Token': access_token,
        'Content-Type': 'application/json'
    })

    for batch_start in range(0, len(orders), batch_size):
        batch_orders = orders[batch_start:batch_start+batch_size]
        
        inventory_items = set()
        for order_edge in batch_orders:
            order = order_edge['node']
            line_items = order.get('lineItems', {}).get('edges', [])
            for item_edge in line_items:
                item = item_edge['node']
                variant = item.get('variant')
                if variant and variant.get('inventoryItem'):
                    inventory_item_id = variant['inventoryItem'].get('id')
                    if inventory_item_id:
                        inventory_items.add(inventory_item_id)
                else:
                    logger.warning(f"Skipping line item in order {order.get('name','')} due to missing variant or inventoryItem.")

        with ThreadPoolExecutor(max_workers=10) as executor:
            list(executor.map(fetch_cost_for_inventory_item, inventory_items))
            list(executor.map(fetch_inventory_levels, inventory_items))
        
        for order_edge in batch_orders:
            order = order_edge['node']
            logger.info(f"Processing Order: {order['name']}")

            country = order['shippingAddress']['country'] if order.get('shippingAddress') else ''
            zone = determine_zone(country)

            order_currency = order['totalPriceSet']['shopMoney']['currencyCode']
            total_value = float(order['totalPriceSet']['shopMoney']['amount'] or 0.0)

            transport_cost = 0.0
            if order.get('totalShippingPriceSet') and order['totalShippingPriceSet'].get('shopMoney'):
                shipping_amount = order['totalShippingPriceSet']['shopMoney'].get('amount')
                if shipping_amount:
                    transport_cost = float(shipping_amount)
            
            line_items = order.get('lineItems', {}).get('edges', [])
            for item_edge in line_items:
                item = item_edge['node']
                variant = item.get('variant')
                if not variant or not variant.get('inventoryItem'):
                    logger.warning(f"Skipping line item in order {order['name']} due to missing variant or inventoryItem.")
                    continue
                
                inventory_item_id = variant['inventoryItem'].get('id')
                if not inventory_item_id:
                    logger.warning(f"No inventory_item_id found for line item in order {order['name']}.")
                    continue

                stock_levels = fetch_inventory_levels(inventory_item_id)
                stock_flores = stock_levels[location_ids["Pr. das Flores, 54"]]
                stock_status = check_stock_status(stock_flores, item['quantity'])
                
                default_selling_price = float(variant.get('price') or 0.0)
                default_compare_price = float(variant.get('compareAtPrice') or 0.0)
                cost_price = fetch_cost_for_inventory_item(inventory_item_id)

                final_price = default_selling_price
                final_compare = default_compare_price
                final_currency = order_currency

                price_diff = final_price - cost_price
                margin_percent = 0.0
                if final_price != 0:
                    margin_percent = 1 - (cost_price / final_price)
                
                variant_weight = float(variant.get('weight', 0.0))

                # Get EAN via variant data
                variant_id = variant.get('id')
                if variant_id:
                    numeric_id = variant_id.split('/')[-1]
                    variant_data = get_variant_data(session, numeric_id)
                    if variant_data and 'variant' in variant_data:
                        ean = variant_data['variant'].get('barcode', '')
                    else:
                        ean = ''
                else:
                    ean = ''

                processed_data.append({
                    "Order Number": order['name'],
                    "Order Date": order['createdAt'],
                    "Country": country,
                    "Currency Code": final_currency,
                    "SKU": item['sku'],
                    "EAN": ean,
                    "Product": item['title'],
                    "Total Value": round(total_value, 2),
                    "Quantity": item['quantity'],
                    "Vendor": variant['product']['vendor'] if variant.get('product') else '',
                    "Stock_Status": stock_status,
                    "Stock_Flores": stock_flores,
                    "Stock_Warehouse": stock_levels[location_ids["Warehouse"]],
                    "Cost_Per_Item": round(cost_price, 2),
                    "Unit_Price": round(final_price, 2),
                    "Compare_Price": round(final_compare, 2),
                    "Price_Diff": round(price_diff, 2),
                    "Margin_Percent": round(margin_percent,4),
                    "Transport_Cost": round(transport_cost, 2),
                    "Numeric_Weight": variant_weight,
                    "Zone": zone,
                    "group": ""
                })
    return processed_data

def convert_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    def margin_to_percent(x):
        if isinstance(x, str) and '%' in x:
            return x
        try:
            x_dec = str(x).replace(',', '.')
            val = float(x_dec) * 100.0
            return f"{val:.2f}%".replace('.', ',')
        except (ValueError, TypeError):
            return x

    def value_to_decimal(x):
        try:
            if isinstance(x, str):
                return float(x.replace(',', '.'))
            return float(x)
        except (ValueError, TypeError):
            return x

    numeric_columns = [
        "Total Value", "Transport Cost", "Cost Per Item", "Unit Price",
        "Compare Price", "Price Diff", "Real Transport Cost", "Real Profit"
    ]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].apply(value_to_decimal)
            df[col] = df[col].apply(lambda x: f"{x:.2f}".replace('.', ',') if isinstance(x, (int, float)) else x)

    if "Margin %" in df.columns:
        df["Margin %"] = df["Margin %"].apply(margin_to_percent)

    return df

def add_total_weight_and_real_cost(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df["Total_Weight"] = ""
        df["Real_Transport_Cost"] = ""
        df["Real_Profit"] = ""
        df["Margin_Percent"] = ""
        return df

    df = df.copy()
    
    def safe_numeric_convert(series):
        if series.dtype == object:
            return pd.to_numeric(series.str.replace(',', '.'), errors='coerce').fillna(0)
        return pd.to_numeric(series, errors='coerce').fillna(0)
    
    df['Total Value'] = safe_numeric_convert(df['Total Value'])
    df['Cost_Per_Item'] = safe_numeric_convert(df['Cost_Per_Item'])
    df['Quantity'] = safe_numeric_convert(df['Quantity'])
    df['Numeric_Weight'] = safe_numeric_convert(df['Numeric_Weight'])
    
    df['Total_Cost'] = df['Cost_Per_Item'] * df['Quantity']
    df['Zone'] = df['Country'].apply(determine_zone)

    order_weights = df.groupby("Order Number")['Numeric_Weight'].transform('sum')
    df["Total_Weight"] = order_weights

    def compute_real_cost(row):
        return get_azul_cost(row['Total_Weight'], row['Zone'])
    
    df["Real_Transport_Cost"] = df.apply(compute_real_cost, axis=1)

    order_totals = df.groupby('Order Number').agg({
        'Total Value': 'first',
        'Total_Cost': 'sum',
        'Real_Transport_Cost': 'first'
    })
    
    order_profits = order_totals['Total Value'] - (order_totals['Total_Cost'] + order_totals['Real_Transport_Cost'])
    order_margin_percent = 1 - (order_profits / order_totals['Total Value'])
    
    df['Real_Profit'] = df['Order Number'].map(order_profits)
    df['Margin_Percent'] = df['Order Number'].map(order_margin_percent)

    df["Total_Weight"] = df["Total_Weight"].apply(lambda x: f"{x:.3f}".replace('.', ','))
    df["Real_Transport_Cost"] = df["Real_Transport_Cost"].apply(lambda x: f"{x:.2f}".replace('.', ','))
    df["Real_Profit"] = df["Real_Profit"].apply(lambda x: f"{x:.2f}".replace('.', ','))
    df["Margin_Percent"] = (df["Margin_Percent"] * 100).round(2).apply(lambda x: f"{x:.2f}%".replace('.', ','))

    df = df.drop(['Total_Cost', 'Zone'], axis=1)

    return df

def split_data_by_group(processed_data: List[Dict]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = pd.DataFrame(processed_data)
    
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    df['Stock_Status'] = df['Stock_Status'].astype(str)
    
    def determine_group(group_df):
        statuses = group_df['Stock_Status'].astype(str)
        if 'OK' in statuses.values:
            if 'OUT OF STOCK' not in statuses.values and 'NOT ENOUGH' not in statuses.values:
                return pd.Series(1, index=group_df.index)
            else:
                return pd.Series(2, index=group_df.index)
        else:
            return pd.Series(3, index=group_df.index)
    
    groups = []
    for _, group_df in df.groupby('Order Number'):
        groups.append(determine_group(group_df))
    df['group'] = pd.concat(groups)
    
    df = add_total_weight_and_real_cost(df)
    
    df_group1 = df[df['group'] == 1].copy()
    df_group2 = df[df['group'] == 2].copy()
    df_group3 = df[df['group'] == 3].copy()
    
    column_names = {
        "Order Number": "Order Number",
        "Order Date": "Order Date",
        "Country": "Country",
        "Currency Code": "Currency Code",
        "SKU": "SKU",
        "EAN": "EAN",
        "Product": "Product",
        "Total Value": "Total Value",
        "Quantity": "Quantity",
        "Vendor": "Vendor",
        "Stock_Status": "Stock Status",
        "Stock_Flores": "Stock Flores",
        "Stock_Warehouse": "Stock Warehouse",
        "Cost_Per_Item": "Cost Per Item",
        "Unit_Price": "Unit Price",
        "Compare_Price": "Compare Price",
        "Price_Diff": "Price Diff",
        "Real_Profit": "Real Profit",
        "Margin_Percent": "Margin %",
        "Transport_Cost": "Transport Cost",
        "Total_Weight": "Total Weight",
        "Real_Transport_Cost": "Real Transport Cost",
        "group": "Group"
    }
    
    df_group1 = df_group1.rename(columns=column_names)
    df_group2 = df_group2.rename(columns=column_names)
    df_group3 = df_group3.rename(columns=column_names)
    
    column_order = [
        "Order Number", "Order Date", "Country", "Currency Code", "SKU", "EAN", "Product", 
        "Total Value", "Quantity", "Vendor", "Stock Status", "Stock Flores", "Stock Warehouse", 
        "Cost Per Item", "Unit Price", "Compare Price", "Price Diff", "Real Profit", 
        "Margin %", "Transport Cost", "Total Weight", "Real Transport Cost", "Group"
    ]

    df_group1 = df_group1[column_order]
    df_group2 = df_group2[column_order]
    df_group3 = df_group3[column_order]
    
    df_group1 = convert_numeric_columns(df_group1)
    df_group2 = convert_numeric_columns(df_group2)
    df_group3 = convert_numeric_columns(df_group3)
    
    return df_group1, df_group2, df_group3

def setup_treeview_style():
    style = ttk.Style()
    style.theme_use("clam")

    style.configure("MyTreeview.Heading",
                    background="#CCCCCC",
                    foreground="black",
                    font=('Arial', 10, 'bold'))

    style.configure("MyTreeview",
                    background="white",
                    foreground="black",
                    font=('Arial', 10),
                    rowheight=25,
                    fieldbackground="white")

    style.layout("MyTreeview", style.layout("Treeview"))
    return style

def configure_columns(tree: ttk.Treeview):
    columns = [
        "Order Number","Order Date","Country","Currency Code","SKU","EAN","Product","Total Value","Quantity","Vendor",
        "Stock_Status","Stock_Flores","Stock_Warehouse","Cost_Per_Item","Unit_Price","Compare_Price","Price_Diff",
        "Margin_Percent","Transport_Cost","Total_Weight","Real_Transport_Cost","Real_Profit","group"
    ]
    tree["columns"] = columns
    tree["show"] = "headings"
    
    col_config = {
        "Order Number": (110, 'center'),
        "Order Date": (180, 'center'),
        "Country": (100, 'center'),
        "Currency Code": (100, 'center'),
        "SKU": (120, 'center'),
        "EAN": (100, 'center'),
        "Product": (300, 'w'),
        "Total Value": (80, 'center'),
        "Quantity": (70, 'center'),
        "Vendor": (100, 'center'),
        "Stock_Status": (100, 'center'),
        "Stock_Flores": (100, 'center'),
        "Stock_Warehouse": (120, 'center'),
        "Cost_Per_Item": (90, 'center'),
        "Unit_Price": (80, 'center'),
        "Compare_Price": (90, 'center'),
        "Price_Diff": (90, 'center'),
        "Margin_Percent": (90, 'center'),
        "Transport_Cost": (90, 'center'),
        "Total_Weight": (90, 'center'),
        "Real_Transport_Cost": (90, 'center'),
        "Real_Profit": (90, 'center'),
        "group": (50, 'center')
    }

    for col in columns:
        width, align = col_config[col]
        anchor = tk.W if align == 'w' else tk.CENTER
        tree.heading(col, text=col, anchor=tk.CENTER)
        tree.column(col, width=width, anchor=anchor)

class CTTCostsViewer:
    def __init__(self, parent):
        self.parent = parent
        self.win = tk.Toplevel(parent)
        self.win.title("CTT Costs")
        self.win.geometry("1200x500")

        self.notebook = ttk.Notebook(self.win)
        self.notebook.pack(expand=True, fill='both')

        self.correo_azul_tab = ttk.Frame(self.notebook, padding=10)
        self.correo_registado_tab = ttk.Frame(self.notebook, padding=10)
        self.country_mapping_tab = ttk.Frame(self.notebook, padding=10)

        self.notebook.add(self.correo_azul_tab, text="Correio Azul")
        self.notebook.add(self.correo_registado_tab, text="Correio Registado")
        self.notebook.add(self.country_mapping_tab, text="Country Mapping")

        self.available_countries = [
            "Portugal", "Angola", "Saudi Arabia", "Australia", "Brazil", "Canada", 
            "Chile", "China", "United Arab Emirates", "France", "Germany", "Ireland",
            "Israel", "Italy", "Japan", "Jordan", "Lithuania", "Morocco", "Mexico",
            "New Zealand", "Poland", "Singapore", "South Africa", "Spain", "Sweden",
            "Switzerland", "Turkey", "United Kingdom", "USA"
        ]
        
        self.transport_methods = ["Correio Azul", "Correio Registado"]
        self.regions = ["USA Zone", "Rest World Zone", "Europe"]

        self.weight_intervals = [
            "< 0.019",
            "[0.020;0.049]",
            "[0.050;0.099]",
            "[0.100;0.249]",
            "[0.250;0.499]",
            "[0.500;0.999]",
            "[1.000;2.000]"
        ]

        self.azul_data = [
            (self.weight_intervals[0], "4,47", "4,89", "5,23"),
            (self.weight_intervals[1], "4,47", "4,89", "5,23"),
            (self.weight_intervals[2], "4,47", "4,89", "5,23"),
            (self.weight_intervals[3], "5,65", "6", "6,66"),
            (self.weight_intervals[4], "7,55", "10,14", "11,75"),
            (self.weight_intervals[5], "10,65", "17,8", "19,8"),
            (self.weight_intervals[6], "16,75", "28", "30,35")
        ]

        self.registado_data = [
            (self.weight_intervals[0], "5,43", "5,79", "6,11"),
            (self.weight_intervals[1], "5,43", "5,79", "6,11"),
            (self.weight_intervals[2], "5,43", "5,79", "6,11"),
            (self.weight_intervals[3], "6,18", "7,4", "8,47"),
            (self.weight_intervals[4], "8,3", "11,55", "12,13"),
            (self.weight_intervals[5], "11,83", "19,15", "21,2"),
            (self.weight_intervals[6], "17,69", "29,5", "31,7")
        ]

        self.azul_tree = self.create_table(self.correo_azul_tab, ["Weight", "Europe", "Rest World", "USA"])
        self.registado_tree = self.create_table(self.correo_registado_tab, ["Weight", "Europe", "Rest World", "USA"])

        self.insert_data(self.azul_tree, self.azul_data)
        self.insert_data(self.registado_tree, self.registado_data)

        country_mapping = [
            ("United States", "Correio Azul", "USA Zone"),
            ("Canada", "Correio Azul", "Rest World Zone"),
            ("Australia", "Correio Azul", "Rest World Zone"),
            ("New Zealand", "Correio Azul", "Rest World Zone"),
            ("France", "Correio Azul", "Europe")
        ]

        self.mapping_controls = ttk.Frame(self.country_mapping_tab)
        self.mapping_controls.pack(fill='x', pady=(0, 10))

        self.country_mapping_tree = self.create_table(self.country_mapping_tab, ["Country of the Order", "Transport Method", "Region"])
        self.insert_data(self.country_mapping_tree, country_mapping)

        self.add_row_button = ttk.Button(self.mapping_controls, text="Add Row", command=self.add_mapping_row)
        self.add_row_button.pack(side='right', padx=5)

        self.current_tree = None
        self.editing_column = None
        self.editing_item = None
        self.entry_widget = None
        self.combo_widget = None

        self.azul_tree.bind("<Double-1>", self.on_double_click)
        self.registado_tree.bind("<Double-1>", self.on_double_click)
        self.country_mapping_tree.bind("<Double-1>", self.on_mapping_double_click)

        self.save_frame = ttk.Frame(self.win)
        self.save_frame.pack(side='bottom', fill='x', pady=10, padx=10, anchor='e')
        self.save_button = ttk.Button(self.save_frame, text="Save", command=self.save_data)
        self.save_button.pack(side='right')

    def add_mapping_row(self):
        self.country_mapping_tree.insert('', 'end', values=("Select Country", "Correio Azul", "Select Region"))
        last_item = self.country_mapping_tree.get_children()[-1]
        self.country_mapping_tree.selection_set(last_item)
        self.country_mapping_tree.focus(last_item)
        self.on_mapping_double_click(None, force_item=last_item)

    def on_mapping_double_click(self, event, force_item=None):
        if force_item:
            item = force_item
        else:
            item = self.country_mapping_tree.selection()[0]
            
        column = self.country_mapping_tree.identify_column(event.x) if event else "#1"
        x, y, w, h = self.country_mapping_tree.bbox(item, column)
        
        values = self.country_mapping_tree.item(item)['values']
        column_id = int(column[1]) - 1
        
        if self.combo_widget:
            self.combo_widget.destroy()
        
        if column_id == 0:
            values_list = self.available_countries
        elif column_id == 1:
            values_list = self.transport_methods
        else:
            values_list = self.regions
            
        self.combo_widget = ttk.Combobox(self.country_mapping_tree, values=values_list)
        self.combo_widget.set(values[column_id])
        self.combo_widget.place(x=x, y=y, width=w, height=h)
        self.combo_widget.focus()
        
        self.combo_widget.bind('<Return>', lambda e: self.finish_mapping_edit(e, item, column_id))
        self.combo_widget.bind('<FocusOut>', lambda e: self.finish_mapping_edit(e, item, column_id))
        
        self.current_tree = self.country_mapping_tree
        self.editing_item = item
        self.editing_column = column_id

    def finish_mapping_edit(self, event, item, column):
        if self.combo_widget:
            new_value = self.combo_widget.get()
            current_values = list(self.country_mapping_tree.item(item)['values'])
            current_values[column] = new_value
            self.country_mapping_tree.item(item, values=current_values)
            self.combo_widget.destroy()
            self.combo_widget = None
            
        self.current_tree = None
        self.editing_item = None
        self.editing_column = None

    def create_table(self, parent_frame, columns):
        tree = ttk.Treeview(parent_frame, columns=columns, show='headings', height=10)
        
        for col in columns:
            tree.column(col, width=200, anchor=tk.CENTER)
            tree.heading(col, text=col)

        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        tree.pack(expand=True, fill='both')
        return tree

    def insert_data(self, tree, data):
        for row in data:
            tree.insert('', 'end', values=row)

    def on_double_click(self, event):
        tree = event.widget
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = tree.identify_column(event.x)
        rowid = tree.identify_row(event.y)
        if not rowid:
            return

        col_index = int(column.replace('#', '')) - 1
        if col_index == 0:
            return

        self.current_tree = tree
        self.editing_item = rowid
        self.editing_column = col_index

        current_value = tree.set(rowid, tree["columns"][col_index])

        x, y, w, h = tree.bbox(rowid, column)
        
        self.entry_widget = ttk.Entry(tree)
        self.entry_widget.insert(0, current_value)
        self.entry_widget.focus()
        self.entry_widget.place(x=x, y=y, width=w, height=h)

        self.entry_widget.bind("<Return>", self.editing_finished)
        self.entry_widget.bind("<FocusOut>", self.editing_finished)

    def editing_finished(self, event):
        if self.entry_widget is None:
            return
        new_value = self.entry_widget.get()
        cols = self.current_tree["columns"]
        self.current_tree.set(self.editing_item, cols[self.editing_column], new_value)
        self.entry_widget.destroy()
        self.entry_widget = None
        self.editing_item = None
        self.editing_column = None
        self.current_tree = None

    def save_data(self):
        print("Correio Azul updated data:")
        for rowid in self.azul_tree.get_children():
            print(self.azul_tree.item(rowid, 'values'))

        print("Correio Registado updated data:")
        for rowid in self.registado_tree.get_children():
            print(self.registado_tree.item(rowid, 'values'))

        print("Country Mapping updated data:")
        for rowid in self.country_mapping_tree.get_children():
            print(self.country_mapping_tree.item(rowid, 'values'))

        messagebox.showinfo("Save", "CTT costs updated successfully!")

class ResultViewer:
    def __init__(self, parent, df):
        self.parent = parent
        self.df = df
        self.win = tk.Toplevel(parent)
        self.win.title("Order Results")
        self.win.geometry("1440x800")  

        self.main_frame = ttk.Frame(self.win, padding=10)  
        self.main_frame.pack(expand=True, fill='both')
        
        self.create_tab_content(self.main_frame, self.df)
        
        style = ttk.Style()
        style.configure('Padded.TButton', padding=(5, 5, 5, 5))
        
        button_frame = ttk.Frame(self.main_frame)
        button_frame.pack(side='bottom', fill='x', padx=5, pady=5)
        
        self.export_button = ttk.Button(button_frame, text="Export to Excel", command=self.export_data, style='Padded.TButton')
        self.export_button.pack(side='right', padx=5, pady=5)

    def create_tab_content(self, tab_frame, df):
        try:
            top_frame = ttk.Frame(tab_frame)
            top_frame.pack(fill=tk.BOTH, expand=True)

            self.top_tree = ttk.Treeview(top_frame, style='MyTreeview')
            
            columns = [
                "Order Number", "Order Date", "Country", "Currency Code", "SKU", "EAN", "Product",
                "Total Value", "Quantity", "Vendor", "Stock Status", "Stock Flores", "Stock Warehouse",
                "Cost Per Item", "Unit Price", "Compare Price", "Price Diff", "Real Profit",
                "Margin %", "Transport Cost", "Total Weight", "Real Transport Cost", "Group"
            ]
            
            self.top_tree["columns"] = columns
            self.top_tree["show"] = "headings"

            column_widths = {
                "Order Number": (120, 'center'),
                "Order Date": (150, 'center'),
                "Country": (100, 'center'),
                "Currency Code": (100, 'center'),
                "SKU": (120, 'center'),
                "EAN": (120, 'center'),
                "Product": (300, 'w'),
                "Total Value": (80, 'center'),
                "Quantity": (70, 'center'),
                "Vendor": (100, 'center'),
                "Stock Status": (100, 'center'),
                "Stock Flores": (80, 'center'),
                "Stock Warehouse": (80, 'center'),
                "Cost Per Item": (80, 'center'),
                "Unit Price": (80, 'center'),
                "Compare Price": (80, 'center'),
                "Price Diff": (80, 'center'),
                "Real Profit": (80, 'center'),
                "Margin %": (80, 'center'),
                "Transport Cost": (80, 'center'),
                "Total Weight": (80, 'center'),
                "Real Transport Cost": (100, 'center'),
                "Group": (80, 'center')
            }

            for col in columns:
                width, anchor = column_widths.get(col, (100, 'center'))
                self.top_tree.heading(col, text=col, anchor=tk.CENTER)
                self.top_tree.column(col, width=width, anchor=anchor)

            distinct_orders = df.drop_duplicates(subset=["Order Number"])[columns].copy()
            for _, row in distinct_orders.iterrows():
                values = [row[col] for col in columns]
                self.top_tree.insert("", tk.END, values=values)

            scrollbar = ttk.Scrollbar(top_frame, orient=tk.VERTICAL, command=self.top_tree.yview)
            self.top_tree.configure(yscrollcommand=scrollbar.set)
            
            self.top_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        except Exception as e:
            logger.error(f"Error creating top content table: {str(e)}")
            traceback.print_exc()
            raise

    def export_data(self):
        messagebox.showerror("Error", "Export from here is disabled. Use main window's export.") 

class OrderAnalysisGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Order Analysis")
        self.root.geometry("1300x900")
        
        self.order_data = []
        self.total_orders = 0
        self.processed_orders = 0
        
        self.start_time = 0
        self.is_running = False
        self.processed_data = []
        self.group_data = None

        self.style = setup_treeview_style()

        self.main_frame = ttk.Frame(self.root, padding=30)
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.create_timer_and_count_widgets()
        self.create_table()
        self.create_progress_section()
        
        style = ttk.Style()
        style.configure('Padded.TButton', padding=(5, 5, 5, 5))

        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.grid(row=3, column=0, columnspan=2, sticky=tk.EW, padx=30, pady=30)

        self.ctt_costs_button = ttk.Button(self.button_frame, text="CTT Costs", command=self.ctt_costs_action, style='Padded.TButton')
        self.start_button = ttk.Button(self.button_frame, text="Start Analysis", command=self.start_analysis, style='Padded.TButton')
        self.export_button = ttk.Button(self.button_frame, text="Export Excel", command=self.export_to_excel_main, style='Padded.TButton')
        self.view_button = ttk.Button(self.button_frame, text="View Orders", command=self.view_orders_action, style='Padded.TButton')
        
        self.ctt_costs_button.pack(side='left', padx=10)
        self.start_button.pack(side='right', padx=10)
        self.export_button.pack(side='right', padx=10)
        self.view_button.pack(side='right', padx=10)

        self.main_frame.columnconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)

    def create_timer_and_count_widgets(self):
        self.timer_frame = ttk.Frame(self.main_frame, padding=30)
        self.timer_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        self.timer_label = ttk.Label(self.timer_frame, text="Time Elapsed:", font=('Arial', 10, 'bold'))
        self.timer_label.grid(row=0, column=0, padx=30)
        
        self.timer_value = ttk.Label(self.timer_frame, text="00:00", font=('Arial', 12))
        self.timer_value.grid(row=0, column=1, padx=30)
        
        self.order_count_label = ttk.Label(self.timer_frame, text="Orders to Process:", font=('Arial', 10, 'bold'))
        self.order_count_label.grid(row=0, column=2, padx=30)
        
        self.order_count_value = ttk.Label(self.timer_frame, text="0/0", font=('Arial', 12))
        self.order_count_value.grid(row=0, column=3, padx=30)

    def create_table(self):
        self.table_frame = ttk.Frame(self.main_frame, padding=30)
        self.table_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.tree = ttk.Treeview(self.table_frame, style='MyTreeview', height=16)
        configure_columns(self.tree)
        
        scrollbar = ttk.Scrollbar(self.table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        self.table_frame.columnconfigure(0, weight=1)

    def create_progress_section(self):
        self.progress_frame = ttk.LabelFrame(self.main_frame, text="Progress", padding=10)
        self.progress_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        ttk.Label(self.progress_frame, text="Fetching Orders:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
        self.fetch_progress_var = tk.DoubleVar()
        self.fetch_progress = ttk.Progressbar(
            self.progress_frame, 
            variable=self.fetch_progress_var,
            maximum=100,
            length=300,
            mode='determinate'
        )
        self.fetch_progress.grid(row=0, column=1, padx=10, pady=10)
        self.fetch_progress_label = ttk.Label(self.progress_frame, text="0%")
        self.fetch_progress_label.grid(row=0, column=2, padx=10, pady=10)
        
        ttk.Label(self.progress_frame, text="Processing Data:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=10)
        self.compute_progress_var = tk.DoubleVar()
        self.compute_progress = ttk.Progressbar(
            self.progress_frame,
            variable=self.compute_progress_var,
            maximum=100,
            length=300,
            mode='determinate'
        )
        self.compute_progress.grid(row=1, column=1, padx=10, pady=10)
        self.compute_progress_label = ttk.Label(self.progress_frame, text="0%")
        self.compute_progress_label.grid(row=1, column=2, padx=10, pady=10)

    def update_timer(self):
        if self.is_running:
            elapsed_time = int(time.time() - self.start_time)
            minutes = elapsed_time // 60
            seconds = elapsed_time % 60
            self.timer_value.config(text=f"{minutes:02d}:{seconds:02d}")
            self.root.after(1000, self.update_timer)

    def update_table(self, order_data: List[Dict]):
        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            for i, row in enumerate(order_data):
                values = []
                for col in self.tree["columns"]:
                    val = row.get(col, "")
                    values.append("" if val is None else str(val))
            
                tag = "oddrow" if i % 2 else "evenrow"
                self.tree.insert("", "end", values=values, tags=(tag,))
        
            self.tree.tag_configure("oddrow", background="#F2D5DB")
            self.tree.tag_configure("evenrow", background="#FAEAED")
        
        except Exception as e:
            logger.error(f"Error in update_table: {str(e)}")
            traceback.print_exc()

    def run_analysis(self):
        try:
            orders = fetch_unfulfilled_orders_with_progress(self)
            logger.info("Processing orders now...")
            self.processed_data = self.process_orders_with_progress(orders)
            
            df_group1, df_group2, df_group3 = split_data_by_group(self.processed_data)
            self.group_data = {1: df_group1, 2: df_group2, 3: df_group3}
            
        except Exception as e:
            err_trace = traceback.format_exc()
            logger.error(f"Error during analysis: {str(e)}\n{err_trace}")
            messagebox.showerror("Error", f"An error occurred during analysis:\n{err_trace}")
        finally:
            self.is_running = False
            self.start_button.config(state='normal')

    def process_orders_with_progress(self, orders: List[Dict]) -> List[Dict]:
        processed_data = []
        batch_size = 10
        total = len(orders)
        
        for batch_idx in range(0, total, batch_size):
            batch_orders = orders[batch_idx:batch_idx+batch_size]
            try:
                batch_processed = process_orders(batch_orders)
            except Exception as e:
                err_trace = traceback.format_exc()
                logger.error(f"Error in process_orders_with_progress: {str(e)}\n{err_trace}")
                messagebox.showerror("Error", f"An error occurred in process_orders_with_progress:\n{err_trace}")
                break
            processed_data.extend(batch_processed)
            
            progress = ((batch_idx + len(batch_orders)) / total) * 100 if total else 0
            self.compute_progress_var.set(progress)
            self.compute_progress_label.config(text=f"{int(progress)}%")
            self.root.update_idletasks()
            
            self.update_table(processed_data[-16:])
        
        return processed_data

    def start_analysis(self):
        self.start_button.config(state='disabled')
        self.start_time = time.time()
        self.is_running = True
        self.update_timer()
        Thread(target=self.run_analysis, daemon=True).start()

    def export_to_excel_main(self):
        if not self.processed_data:
            messagebox.showerror("Error", "No data available to export. Please run the analysis first.")
            return
        
        try:
            df_group1, df_group2, df_group3 = split_data_by_group(self.processed_data)
            timestamp_str = datetime.now().strftime('%Y_%m_%d_%M_%S')
            file_name = f"Cosmetyque_Orders_Grouped_{timestamp_str}.xlsx"
            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Analysis Report",
                initialfile=file_name
            )
            
            if file_path:
                save_excel_file(file_path, df_group1, df_group2, df_group3)
        except Exception as e:
            err_trace = traceback.format_exc()
            logger.error(f"Error in export_to_excel_main: {str(e)}\n{err_trace}")
            messagebox.showerror("Error", f"An error occurred in export_to_excel_main:\n{err_trace}")

    def view_orders_action(self):
        if self.group_data is not None:
            ResultViewer(self.root, self.group_data)
        else:
            messagebox.showerror("Error", "No order data available. Please run the analysis first.")

    def ctt_costs_action(self):
        CTTCostsViewer(self.root)

def main():
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(('localhost', 12345))
        root = tk.Tk()
        app = OrderAnalysisGUI(root)
        root.mainloop()
    except socket.error:
        messagebox.showwarning("Warning", "Application is already running!")
    finally:
        sock.close()

if __name__ == "__main__":
    main()
