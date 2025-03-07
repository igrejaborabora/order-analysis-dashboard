import os
import logging
import requests
import pandas as pd
import time
from typing import Dict, List, Optional, Any, Tuple
from datetime import datetime
from tenacity import retry, stop_after_attempt, wait_exponential
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from threading import Thread
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import re
from functools import lru_cache
from concurrent.futures import ThreadPoolExecutor
import traceback
import subprocess
import urllib3

# Disable SSL warnings
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Setup logging
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('orders_analysis.log'),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Load environment variables
load_dotenv()

# Shopify API Configuration following best practices
SHOPIFY_STORE_URL = os.getenv('COSMETYQUE_SHOPIFY_STORE_URL', 'cosmetyque.myshopify.com')
SHOPIFY_ACCESS_TOKEN = os.getenv('COSMETYQUE_ACCESS_TOKEN')
SHOPIFY_API_VERSION = os.getenv('COSMETYQUE_API_VERSION', '2024-01')
SHOPIFY_STOREFRONT_TOKEN = os.getenv('COSMETYQUE_STOREFRONT_TOKEN')

# Location IDs
LOCATION_ID_FLORES = os.getenv('COSMETYQUE_LOCATION_ID_FLORES')
LOCATION_ID_WAREHOUSE = os.getenv('COSMETYQUE_LOCATION_ID_WAREHOUSE')

# Validate required environment variables
required_vars = {
    'SHOPIFY_ACCESS_TOKEN': SHOPIFY_ACCESS_TOKEN,
    'LOCATION_ID_FLORES': LOCATION_ID_FLORES,
    'LOCATION_ID_WAREHOUSE': LOCATION_ID_WAREHOUSE
}

missing_vars = [var for var, value in required_vars.items() if not value]
if missing_vars:
    error_msg = f"Missing required environment variables: {', '.join(missing_vars)}"
    logger.error(error_msg)
    messagebox.showerror("Error", error_msg)
    raise SystemExit(error_msg)

if not SHOPIFY_STOREFRONT_TOKEN:
    logger.warning("COSMETYQUE_STOREFRONT_TOKEN not found. International pricing might be limited.")

# API endpoints
SHOPIFY_GRAPHQL_URL = f"https://{SHOPIFY_STORE_URL}/admin/api/{SHOPIFY_API_VERSION}/graphql.json"
SHOPIFY_REST_URL = f"https://{SHOPIFY_STORE_URL}/admin/api/{SHOPIFY_API_VERSION}"
SHOPIFY_STOREFRONT_URL = f"https://{SHOPIFY_STORE_URL}/api/{SHOPIFY_API_VERSION}/storefront/graphql.json"

# Location mapping
LOCATION_IDS = {
    "Pr. das Flores, 54": LOCATION_ID_FLORES,
    "Warehouse": LOCATION_ID_WAREHOUSE
}

# Initialize global caches
inventory_cache = {}
cost_cache: Dict[str, float] = {}

@lru_cache(maxsize=1000)
def fetch_cost_for_inventory_item(inventory_item_id: str) -> float:
    if inventory_item_id in cost_cache:
        return cost_cache[inventory_item_id]
        
    inventory_item_id_numeric = re.search(r'\d+$', inventory_item_id).group()
    cost_url = f"{SHOPIFY_REST_URL}/inventory_items/{inventory_item_id_numeric}.json"
    
    headers = {
        "Content-Type": "application/json",
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN
    }
    
    while True:
        response = requests.get(cost_url, headers=headers)
        
        if response.status_code == 429:
            logger.warning("Rate limit exceeded. Retrying in 1 second...")
            time.sleep(1)
            continue
        
        response.raise_for_status()
        
        data = response.json()
        cost = float(data.get("inventory_item", {}).get("cost", 0.0))
        cost_cache[inventory_item_id] = cost
        return cost

@lru_cache(maxsize=1000)
def fetch_inventory_levels(inventory_item_id: str) -> Dict[str, int]:
    global inventory_cache  # Ensure we're using the global cache
    
    inventory_query = """
    query($inventoryItemId: ID!) {
      inventoryItem(id: $inventoryItemId) {
        inventoryLevels(first: 10) {
          edges {
            node {
              quantities(names: ["available"]) {
                name
                quantity
              }
              location {
                id
                name
              }
            }
          }
        }
      }
    }
    """

    if inventory_item_id in inventory_cache:
        logger.info(f"Using cached inventory data for {inventory_item_id}")
        return inventory_cache[inventory_item_id]

    # Default empty stock levels in case of failure
    stock_levels = {location: 0 for location in LOCATION_IDS.values()}
    
    try:
        logger.info(f"Fetching inventory levels for item: {inventory_item_id}")
        response = execute_graphql_query(inventory_query, {"inventoryItemId": inventory_item_id})
        
        # Add extensive debugging
        logger.debug(f"Inventory API response: {response}")
        logger.debug(f"Expected location IDs: {LOCATION_IDS}")
        
        # Check if response contains the expected data structure
        if not isinstance(response, dict):
            logger.error(f"Response is not a dictionary: {response}")
            inventory_cache[inventory_item_id] = stock_levels
            return stock_levels
            
        if 'data' not in response:
            logger.warning(f"Missing 'data' key in GraphQL response for inventory item {inventory_item_id}")
            
            # Check if there are errors in the response
            if 'errors' in response:
                for error in response['errors']:
                    logger.error(f"GraphQL error: {error.get('message', 'Unknown error')}")
                    
                    # Handle rate limiting specifically (Shopify's 2 requests/second limit)
                    if 'throttled' in str(error.get('message', '')).lower() or 'rate limit' in str(error.get('message', '')).lower():
                        logger.warning("Rate limit detected, sleeping before retry")
                        time.sleep(1)  # Sleep for 1 second before next request
            
            # Return default empty stock levels
            inventory_cache[inventory_item_id] = stock_levels
            return stock_levels
        
        # Check if the inventoryItem exists
        if response['data'].get('inventoryItem') is None:
            logger.warning(f"Inventory item {inventory_item_id} not found")
            inventory_cache[inventory_item_id] = stock_levels
            return stock_levels
            
        inventory_levels = response['data']['inventoryItem']['inventoryLevels']['edges']
        
        logger.info(f"Found {len(inventory_levels)} inventory level entries for item {inventory_item_id}")
        
        # Clean location matching with detailed logging
        for level in inventory_levels:
            node = level['node']
            location_id = node['location']['id']
            location_name = node['location']['name']
            
            # Extract available quantity from quantities array
            available = 0
            for quantity_entry in node['quantities']:
                if quantity_entry['name'] == 'available':
                    available = int(quantity_entry['quantity'])
                    break
            
            logger.info(f"Location from API: ID={location_id}, Name={location_name}, Available={available}")
            
            # Try to match location by ID first
            if location_id in stock_levels:
                stock_levels[location_id] = available
                logger.info(f"Matched location by ID: {location_id} with {available} units")
            # Then try to match by name using our mapping
            else:
                # Look for matching location by name
                matched = False
                for loc_name, loc_id in LOCATION_IDS.items():
                    if loc_name.lower() in location_name.lower() or location_name.lower() in loc_name.lower():
                        stock_levels[loc_id] = available
                        logger.info(f"Matched location by name: {location_name} -> {loc_name} with {available} units")
                        matched = True
                        break
                
                if not matched:
                    logger.warning(f"Could not match location {location_name} ({location_id}) to any known location")
                
        logger.info(f"Final stock levels for {inventory_item_id}: {stock_levels}")
        inventory_cache[inventory_item_id] = stock_levels
        return stock_levels
        
    except Exception as e:
        logger.error(f"Error fetching inventory levels for {inventory_item_id}: {str(e)}")
        logger.error(f"Error type: {type(e).__name__}")
        logger.error(traceback.format_exc())
        inventory_cache[inventory_item_id] = stock_levels
        return stock_levels

@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=4, max=10))
def execute_graphql_query(query: str, variables: Optional[Dict] = None) -> Dict[str, Any]:
    url = SHOPIFY_GRAPHQL_URL
    headers = {
        "X-Shopify-Access-Token": SHOPIFY_ACCESS_TOKEN,
        "Content-Type": "application/json",
    }
    
    try:
        response = requests.post(
            url,
            json={"query": query, "variables": variables},
            headers=headers,
            verify=True
        )
        response.raise_for_status()
        return response.json()
    except requests.exceptions.SSLError:
        logger.warning("SSL verification failed, attempting without verification (not recommended for production)")
        response = requests.post(
            url,
            json={"query": query, "variables": variables},
            headers=headers,
            verify=False
        )
        response.raise_for_status()
        return response.json()

# Including barcode and removing currency code and override from final output
orders_query = """
query($cursor: String) {
  orders(first: 50, after: $cursor, query: "fulfillment_status:unfulfilled AND status:open", sortKey: CREATED_AT, reverse: true) {
    edges {
      cursor
      node {
        id
        name
        createdAt
        customer {
          firstName
          lastName
        }
        shippingAddress {
          country
        }
        totalPriceSet {
          shopMoney {
            amount
            currencyCode
          }
          presentmentMoney {
            amount
            currencyCode
          }
        }
        totalShippingPriceSet {
          shopMoney {
            amount
            currencyCode
          }
        }
        displayFulfillmentStatus
        lineItems(first: 50) {
          edges {
            node {
              title
              quantity
              sku
              variant {
                id
                barcode
                inventoryItem {
                  id
                }
                price
                compareAtPrice
                product {
                  vendor
                }
                weight
                weightUnit
              }
              originalUnitPriceSet {
                presentmentMoney {
                  amount
                  currencyCode
                }
              }
            }
          }
        }
      }
    }
    pageInfo {
      hasNextPage
      endCursor
    }
  }
}
"""

count_query = """
query {
  orders(first: 250, query: "fulfillment_status:unfulfilled AND status:open") {
    pageInfo {
      hasNextPage
      endCursor
    }
    edges {
      node {
        id
      }
    }
  }
}
"""

def fetch_unfulfilled_orders_with_progress(gui_app) -> List[Dict]:
    all_orders = []
    result = execute_graphql_query(count_query)
    if not result or 'data' not in result or 'orders' not in result['data']:
        error_msg = "Invalid response structure from GraphQL API"
        logger.error(f"{error_msg}. Response: {result}")
        raise Exception(error_msg)
    
    initial_orders = result['data']['orders']['edges']
    total_count = len(initial_orders)
    has_more = result['data']['orders']['pageInfo']['hasNextPage']
    next_cursor = result['data']['orders']['pageInfo']['endCursor']
    
    while has_more:
        count_with_cursor = f"""
        query {{
            orders(first: 250, after: "{next_cursor}", query: "fulfillment_status:unfulfilled AND status:open") {{
                pageInfo {{
                    hasNextPage
                    endCursor
                }}
                edges {{
                    node {{
                        id
                    }}
                }}
            }}
        }}
        """
        result = execute_graphql_query(count_with_cursor)
        if not result or 'data' not in result or 'orders' not in result['data']:
            break
        batch = result['data']['orders']['edges']
        total_count += len(batch)
        has_more = result['data']['orders']['pageInfo']['hasNextPage']
        next_cursor = result['data']['orders']['pageInfo']['endCursor']
    
    gui_app.total_orders = total_count
    gui_app.order_count_value.config(text=f"0/{gui_app.total_orders}")
    
    cursor = None
    while True:
        variables = {"cursor": cursor} if cursor else {}
        logger.info(f"Fetching orders with cursor: {cursor}")
        result = execute_graphql_query(orders_query, variables)
        
        if not result or 'data' not in result or 'orders' not in result['data']:
            break
        
        orders_data = result['data']['orders']
        if not orders_data.get('edges'):
            break
        
        all_orders.extend(orders_data['edges'])
        
        gui_app.processed_orders = len(all_orders)
        gui_app.order_count_value.config(text=f"{gui_app.processed_orders}/{gui_app.total_orders}")
        progress = (len(all_orders) / total_count * 100) if total_count > 0 else 0
        gui_app.fetch_progress_var.set(progress)
        gui_app.fetch_progress_label.config(text=f"{int(progress)}%")
        gui_app.root.update_idletasks()
        
        if not orders_data['pageInfo']['hasNextPage']:
            break
        cursor = orders_data['pageInfo']['endCursor']
        time.sleep(0.5)
    
    return all_orders

def check_stock_status(stock: int, required_quantity: int) -> str:
    """
    Determine the stock status based on available stock and required quantity.
    
    Args:
        stock (int): Total available stock across all locations
        required_quantity (int): Quantity needed for the order
        
    Returns:
        str: Stock status - "OK", "NOT ENOUGH", or "OUT OF STOCK"
    """
    if stock == 0:
        return "OUT OF STOCK"
    elif stock < required_quantity:
        return "NOT ENOUGH"
    else:
        return "OK"

def format_excel_output(writer: pd.ExcelWriter, df: pd.DataFrame, sheet_name: str) -> None:
    worksheet = writer.sheets[sheet_name]
    
    for idx, column in enumerate(df.columns, 1):
        col_letter = get_column_letter(idx)
        header_cell = worksheet.cell(row=1, column=idx)
        header_cell.font = Font(bold=True)
        header_cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        max_length = max(len(str((worksheet.cell(row=r, column=idx).value) or '')) for r in range(1, len(df)+2))
        adjusted_width = min(max_length + 2, 50)
        worksheet.column_dimensions[col_letter].width = adjusted_width

def apply_excel_formatting(workbook):
    header_color = 'CE5870'
    colors = ['F2D5DB', 'FAEAED']

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        
        worksheet.row_dimensions[1].height = 50
        
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = PatternFill(start_color=header_color, end_color=header_color, fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        current_order = None
        color_index = 0
        
        for row in range(2, worksheet.max_row + 1):
            order_number = worksheet.cell(row=row, column=1).value
            if order_number != current_order:
                current_order = order_number
                color_index = (color_index + 1) % 2
            
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.fill = PatternFill(start_color=colors[color_index], end_color=colors[color_index], fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                header_cell = worksheet.cell(row=1, column=col)
                if header_cell.value == 'Margin %':
                    if cell.value is not None and row > 1:
                        try:
                            value = float(str(cell.value).replace('%', ''))
                            cell.value = value
                            cell.number_format = '0.00%'
                        except (ValueError, TypeError):
                            pass
        
        for col in range(1, worksheet.max_column + 1):
            max_length = 0
            column = get_column_letter(col)
            for r in range(1, worksheet.max_row + 1):
                cell = worksheet.cell(row=r, column=col)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = max_length + 2
            worksheet.column_dimensions[column].width = adjusted_width

def save_excel_file(file_path: str, df_group1: pd.DataFrame, df_group2: pd.DataFrame, df_group3: pd.DataFrame) -> None:
    try:
        writer = pd.ExcelWriter(file_path, engine='openpyxl')
        
        df_group1.to_excel(writer, sheet_name='Group 1 - Full Stock', index=False)
        df_group2.to_excel(writer, sheet_name='Group 2 - Partial Stock', index=False)
        df_group3.to_excel(writer, sheet_name='Group 3 - No Stock', index=False)
        
        workbook = writer.book
        apply_excel_formatting(workbook)
        writer.close()
        
        logger.info(f"Excel file saved successfully: {file_path}")
        messagebox.showinfo("Success", "Analysis report saved successfully!")
        os.startfile(os.path.abspath(file_path))
    except Exception as e:
        err_trace = traceback.format_exc()
        logger.error(f"Error saving Excel file: {str(e)}\n{err_trace}")
        messagebox.showerror("Error", f"An error occurred while saving the file: {str(e)}")
        raise

def determine_zone(country: str) -> str:
    country = country.strip().lower()
    if country == 'portugal':
        return 'Europe'
    elif country == 'united states':
        return 'USA'
    elif country == 'france':
        return 'Europe'
    elif country in ['canada', 'australia', 'new zealand']:
        return 'RestWorld'
    else:
        return 'RestWorld'

def get_azul_cost(total_weight_kg: float, zone: str) -> float:
    adjusted_weight = total_weight_kg * 1.10
    intervals = [
        (0.019, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.049, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.099, {"Europe":4.47,"RestWorld":4.89,"USA":5.23}),
        (0.249, {"Europe":5.65,"RestWorld":6.00,"USA":6.66}),
        (0.499, {"Europe":7.55,"RestWorld":10.14,"USA":11.75}),
        (0.999, {"Europe":10.65,"RestWorld":17.8,"USA":19.8}),
        (2.000, {"Europe":16.75,"RestWorld":28.00,"USA":30.35})
    ]

    for ub, costs in intervals:
        if adjusted_weight <= ub:
            return costs.get(zone, costs["RestWorld"])
    last = intervals[-1][1]
    return last.get(zone, last["RestWorld"])

def get_ctt_expresso_cost(total_weight_kg: float) -> float:
    # CTT Expresso prices for Portugal
    # Até 1Kg	3.74 €
    # Até 5Kg	4.03 €
    # Até 10Kg	4.43 €
    # Até 15Kg	4.81 €
    # Até 20Kg	5.26 €
    # Até 25Kg	5.99 €
    # Até 30Kg	6.54 €
    # Mais de 30Kg 200 €
    if total_weight_kg <= 1:
        return 3.74
    elif total_weight_kg <= 5:
        return 4.03
    elif total_weight_kg <= 10:
        return 4.43
    elif total_weight_kg <= 15:
        return 4.81
    elif total_weight_kg <= 20:
        return 5.26
    elif total_weight_kg <= 25:
        return 5.99
    elif total_weight_kg <= 30:
        return 6.54
    else:
        return 200.00

def get_variant_pricing(variant_id: str) -> Dict[str, Any]:
    query = """
    query($id: ID!) {
      productVariant(id: $id) {
        id
        price
        compareAtPrice
      }
    }
    """
    data = execute_graphql_query(query, {"id": variant_id})
    variant_data = data.get("data", {}).get("productVariant", {})
    
    base_price = float(variant_data.get("price", 0.0))
    base_compare = variant_data.get("compareAtPrice")
    base_compare = float(base_compare) if base_compare else 0.0

    return {
        "base_price": base_price,
        "base_compare": base_compare
    }

def process_orders(orders: List[Dict]) -> List[Dict]:
    processed_data = []
    inventory_items = set()
    
    for batch_orders in orders:
        for order_edge in batch_orders['edges']:
            order = order_edge['node']
            for item_edge in order['lineItems']['edges']:
                item = item_edge['node']
                if item.get('variant') and item['variant'].get('inventoryItem'):
                    inventory_items.add(item['variant']['inventoryItem']['id'])

        with ThreadPoolExecutor(max_workers=10) as executor:
            list(executor.map(fetch_cost_for_inventory_item, inventory_items))
            list(executor.map(fetch_inventory_levels, inventory_items))

        for order_edge in batch_orders['edges']:
            order = order_edge['node']
            logger.info(f"Processing Order: {order['name']}")

            country = order['shippingAddress']['country'] if order.get('shippingAddress') else ''
            zone = determine_zone(country)

            total_value_eur = float(order['totalPriceSet']['shopMoney']['amount'])
            total_value_local = float(order['totalPriceSet']['presentmentMoney']['amount'])
            local_currency_code = order['totalPriceSet']['presentmentMoney']['currencyCode']

            # We no longer need Currency Code and Override Price columns in final output
            order_currency = order['totalPriceSet']['shopMoney']['currencyCode']

            charged_transport_cost = 0.0
            if order.get('totalShippingPriceSet') and order['totalShippingPriceSet'].get('shopMoney'):
                shipping_amount = order['totalShippingPriceSet']['shopMoney'].get('amount')
                if shipping_amount:
                    charged_transport_cost = float(shipping_amount)

            total_weight = 0.0
            logger.info(f"\n=== Weight Calculation for Order {order['name']} ===")
            for item_edge in order['lineItems']['edges']:
                try:
                    item = item_edge['node']
                    if not item.get('variant'):
                        continue
                    variant_weight = float(item['variant'].get('weight', 0.0))
                    quantity = item.get('quantity', 1)
                    total_weight += variant_weight * quantity
                    logger.info(f"Product: {item.get('title')} - Weight: {variant_weight:.3f} kg")
                except Exception as e:
                    logger.error(f"Error calculating weight for item in order {order['name']}: {str(e)}")
                    continue
            logger.info(f"Total Order Weight: {total_weight:.3f} kg")

            total_order_cost = 0.0
            for cost_item_edge in order['lineItems']['edges']:
                if not cost_item_edge['node'].get('variant') or not cost_item_edge['node']['variant'].get('inventoryItem'):
                    continue
                item_node = cost_item_edge['node']
                item_inventory_id = item_node['variant']['inventoryItem']['id']
                item_quantity = int(item_node['quantity'])
                item_cost = fetch_cost_for_inventory_item(item_inventory_id)
                total_order_cost += item_cost * item_quantity

            # If country is Portugal, use CTT Expresso cost:
            if country.strip().lower() == 'portugal':
                real_transport_cost = get_ctt_expresso_cost(total_weight)
            else:
                real_transport_cost = get_azul_cost(total_weight, zone)

            real_profit = total_value_eur - total_order_cost - real_transport_cost
            margin_percent = (real_profit / total_value_eur) * 100 if total_value_eur > 0 else 0

            logger.info(f"\n=== Margin Calculation for Order {order['name']} ===")
            logger.info(f"Total Value (EUR): {total_value_eur:.2f}")
            logger.info(f"Total Cost of All Products: {total_order_cost:.2f}")
            logger.info(f"Real Transport Cost: {real_transport_cost:.2f}")
            logger.info(f"Real Profit: {real_profit:.2f}")
            logger.info(f"Margin %: {margin_percent:.2f}%")

            total_value_local_str = f"{total_value_local:.2f}".replace('.', ',')
            total_local_value_display = f"{local_currency_code} {total_value_local_str}"
            XE = total_value_local / total_value_eur if total_value_eur != 0 else 0.0
            XE_str = f"{XE:.9f}".replace('.', ',')

            for item_edge in order['lineItems']['edges']:
                try:
                    item = item_edge['node']
                    if not item.get('variant') or not item['variant'].get('inventoryItem'):
                        logger.warning(f"Skipping item in order {order['name']} due to missing variant or inventory information")
                        continue

                    inventory_item_id = item['variant']['inventoryItem']['id']
                    stock_levels = fetch_inventory_levels(inventory_item_id)
                    stock_flores = stock_levels.get(LOCATION_IDS["Pr. das Flores, 54"], 0)
                    stock_warehouse = stock_levels.get(LOCATION_IDS["Warehouse"], 0)
                    total_stock = stock_flores + stock_warehouse
                    
                    quantity = int(item['quantity'])
                    # Use total stock from both locations for status determination
                    stock_status = check_stock_status(total_stock, quantity)
                    
                    # Add debug logging
                    logger.info(f"Item in order {order['name']}: Stock Flores={stock_flores}, Warehouse={stock_warehouse}, Total={total_stock}, Required={quantity}, Status={stock_status}")
                    
                    cost_per_item = fetch_cost_for_inventory_item(inventory_item_id)

                    # Base EUR pricing
                    variant_id = item['variant']['id']
                    pricing = get_variant_pricing(variant_id)
                    
                    base_price = pricing["base_price"]
                    base_compare = pricing["base_compare"]

                    # Fetch barcode (EAN)
                    ean = item['variant'].get('barcode', '')

                    # Fetch the local item price directly from originalUnitPriceSet
                    local_item_amount = float(item['originalUnitPriceSet']['presentmentMoney']['amount'])
                    local_item_currency = item['originalUnitPriceSet']['presentmentMoney']['currencyCode']

                    # Format Item Local Price
                    item_local_price_str = f"{local_item_amount:.2f}".replace('.', ',')
                    item_local_price_display = f"{local_item_currency} {item_local_price_str}"

                    # Price Diff Calculation:
                    # Price Diff = (Item Local Price / XE) - Shopify Cost
                    local_item_in_eur = local_item_amount / XE if XE != 0 else 0.0
                    price_diff = local_item_in_eur - cost_per_item
                    price_diff_str = f"€ {round(price_diff, 2):,.2f}".replace('.', ',')

                    processed_data.append({
                        "Order Number": order['name'],
                        "Order Date": order['createdAt'],
                        "Country": country,
                        "SKU": item.get('sku', ''),
                        "EAN": ean,
                        "Product": item.get('title', 'Unknown Product'),
                        "Quantity": quantity,
                        "Vendor": item.get('variant', {}).get('product', {}).get('vendor', 'Unknown'),
                        "Shopify Cost": f"€ {round(cost_per_item, 2):,.2f}".replace('.', ','),
                        "Base Price (€)": f"€ {round(base_price, 2):,.2f}".replace('.', ','),
                        "Compare At Price (€)": f"€ {round(base_compare, 2):,.2f}".replace('.', ','),
                        "Item Local Price": item_local_price_display,
                        "Total Value": f"€ {round(total_value_eur, 2):,.2f}".replace('.', ','),
                        "Total Local Value": total_local_value_display,
                        "XE": XE_str,
                        "Price Diff": price_diff_str,
                        "Real Profit": f"€ {round(real_profit, 2):,.2f}".replace('.', ','),
                        "Margin %": f"{round(margin_percent, 2):,.2f} %".replace('.', ','),
                        "Stock Status": stock_status,
                        "Stock Flores": stock_flores,
                        "Stock Warehouse": stock_warehouse,
                        "Numeric_Weight": float(item['variant'].get('weight', 0.0)),
                        "Total Weight": "",
                        "Transport Cost": f"€ {round(charged_transport_cost, 2):,.2f}".replace('.', ','),
                        "Real Transport Cost": f"€ {round(real_transport_cost, 2):,.2f}".replace('.', ','),
                        "group": ""
                    })

                except Exception as e:
                    logger.error(f"Error processing item in order {order['name']}: {str(e)}")
                    continue

    return processed_data

def convert_numeric_columns(df: pd.DataFrame) -> pd.DataFrame:
    def margin_to_percent(x):
        if isinstance(x, str) and '%' in x:
            return x
        try:
            x_dec = str(x).replace(',', '.')
            val = float(x_dec) * 100.0
            return f"{val:.2f}%".replace('.', ',')
        except (ValueError, TypeError):
            return x

    def value_to_decimal(x):
        try:
            if isinstance(x, str):
                return float(x.replace(',', '.'))
            return float(x)
        except (ValueError, TypeError):
            return x

    numeric_columns = [
        "Total Value", "Transport Cost", "Shopify Cost", "Base Price (€)",
        "Compare At Price (€)", "Price Diff", "Real Profit", "Margin %"
    ]

    for col in numeric_columns:
        if col in df.columns:
            df[col] = df[col].apply(value_to_decimal)
            df[col] = df[col].apply(lambda v: f"{v:.2f}".replace('.', ',') if isinstance(v, (int, float)) else v)

    if "Margin %" in df.columns:
        df["Margin %"] = df["Margin %"].apply(margin_to_percent)

    return df

def add_total_weight_and_real_cost(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        df["Total Weight"] = ""
        return df

    df = df.copy()
    
    def safe_numeric_convert(series):
        if series.dtype == object:
            return pd.to_numeric(series.str.replace(',', '.'), errors='coerce').fillna(0)
        return pd.to_numeric(series, errors='coerce').fillna(0)
    
    df['Numeric_Weight'] = safe_numeric_convert(df['Numeric_Weight'])
    
    order_weights = df.groupby("Order Number")['Numeric_Weight'].transform('sum')
    df["Total Weight"] = order_weights
    df["Total Weight"] = df["Total Weight"].apply(lambda x: f"{x:.3f}".replace('.', ','))
    
    return df

def split_data_by_group(processed_data: List[Dict]) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    df = pd.DataFrame(processed_data)
    
    if df.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    
    # Debug stock status values
    logger.info(f"Stock Status values in dataset: {df['Stock Status'].unique()}")
    df['Stock Status'] = df['Stock Status'].astype(str)
    
    def determine_group(group_df):
        # Count occurrences of each status
        status_counts = group_df['Stock Status'].value_counts()
        logger.info(f"Order {group_df['Order Number'].iloc[0]} has stock statuses: {status_counts}")
        
        # Group 1: All items are "OK"
        if len(status_counts) == 1 and 'OK' in status_counts:
            logger.info(f"Order {group_df['Order Number'].iloc[0]} assigned to Group 1")
            return pd.Series(1, index=group_df.index)
        
        # Group 3: Any item is "OUT OF STOCK"
        if 'OUT OF STOCK' in status_counts:
            logger.info(f"Order {group_df['Order Number'].iloc[0]} assigned to Group 3")
            return pd.Series(3, index=group_df.index)
        
        # Group 2: Mix of "OK" and "NOT ENOUGH" or all "NOT ENOUGH"
        logger.info(f"Order {group_df['Order Number'].iloc[0]} assigned to Group 2")
        return pd.Series(2, index=group_df.index)
    
    groups = []
    for _, group_df in df.groupby('Order Number'):
        groups.append(determine_group(group_df))
    df['group'] = pd.concat(groups)
    
    # Debug final group assignments
    for group_num in [1, 2, 3]:
        order_count = len(df[df['group'] == group_num]['Order Number'].unique())
        logger.info(f"Total orders in Group {group_num}: {order_count}")
    
    df = add_total_weight_and_real_cost(df)
    
    column_names = {
        "Order Number": "Order Number",
        "Order Date": "Order Date",
        "Country": "Country",
        "SKU": "SKU",
        "EAN": "EAN",
        "Product": "Product",
        "Quantity": "Quantity",
        "Vendor": "Vendor",
        "Shopify Cost": "Shopify Cost",
        "Base Price (€)": "Base Price (€)",
        "Compare At Price (€)": "Compare At Price (€)",
        "Item Local Price": "Item Local Price",
        "Total Value": "Total Value",
        "Total Local Value": "Total Local Value",
        "XE": "XE",
        "Price Diff": "Price Diff",
        "Real Profit": "Real Profit",
        "Margin %": "Margin %",
        "Stock Status": "Stock Status",
        "Stock Flores": "Stock Flores",
        "Stock Warehouse": "Stock Warehouse",
        "Total Weight": "Total Weight",
        "Transport Cost": "Transport Cost",
        "Real Transport Cost": "Real Transport Cost",
        "group": "Group"
    }
    
    df = df.rename(columns=column_names)
    
    # Removed Currency Code and Override Price from the final output
    column_order = [
        "Order Number", "Order Date", "Country", "SKU", "EAN", "Product",
        "Quantity", "Vendor", "Shopify Cost", "Base Price (€)",
        "Compare At Price (€)", "Item Local Price", "Total Value", "Total Local Value", "XE", "Price Diff",
        "Real Profit", "Margin %", "Stock Status", "Stock Flores", "Stock Warehouse",
        "Total Weight", "Transport Cost", "Real Transport Cost", "Group"
    ]

    missing_cols = [c for c in column_order if c not in df.columns]
    for c in missing_cols:
        df[c] = ""
    
    df = df[column_order]
    
    df_group1 = df[df['Group'] == 1].copy()
    df_group2 = df[df['Group'] == 2].copy()
    df_group3 = df[df['Group'] == 3].copy()
    
    df_group1 = convert_numeric_columns(df_group1)
    df_group2 = convert_numeric_columns(df_group2)
    df_group3 = convert_numeric_columns(df_group3)
    
    return df_group1, df_group2, df_group3

def setup_treeview_style():
    style = ttk.Style()
    style.theme_use("clam")

    style.configure("MyTreeview.Heading",
                    background="#CCCCCC",
                    foreground="black",
                    font=('Arial', 10, 'bold'))

    style.configure("MyTreeview",
                    background="white",
                    foreground="black",
                    font=('Arial', 10),
                    rowheight=25,
                    fieldbackground="white")

    style.layout("MyTreeview", style.layout("Treeview"))
    return style

def configure_columns(tree: ttk.Treeview):
    columns = [
        "Order Number","Order Date","Country","SKU","EAN","Product","Quantity","Vendor",
        "Shopify Cost","Base Price (€)","Compare At Price (€)","Item Local Price","Total Value","Total Local Value","XE","Price Diff","Real Profit","Margin %","Stock Status","Stock Flores","Stock Warehouse","Total Weight","Transport Cost","Real Transport Cost","Group"
    ]
    tree["columns"] = columns
    tree["show"] = "headings"
    
    col_config = {
        "Order Number": (110, 'center'),
        "Order Date": (180, 'center'),
        "Country": (100, 'center'),
        "SKU": (120, 'center'),
        "EAN": (100, 'center'),
        "Product": (250, 'center'),
        "Quantity": (70, 'center'),
        "Vendor": (100, 'center'),
        "Shopify Cost": (90, 'center'),
        "Base Price (€)": (90, 'center'),
        "Compare At Price (€)": (110, 'center'),
        "Item Local Price": (130, 'center'),
        "Total Value": (80, 'center'),
        "Total Local Value": (130, 'center'),
        "XE": (100, 'center'),
        "Price Diff": (70, 'center'),
        "Real Profit": (70, 'center'),
        "Margin %": (70, 'center'),
        "Stock Status": (90, 'center'),
        "Stock Flores": (90, 'center'),
        "Stock Warehouse": (100, 'center'),
        "Total Weight": (90, 'center'),
        "Transport Cost": (90, 'center'),
        "Real Transport Cost": (100, 'center'),
        "Group": (50, 'center')
    }

    for col in columns:
        width, align = col_config[col]
        anchor = tk.CENTER if align == 'center' else tk.W
        tree.heading(col, text=col, anchor=tk.CENTER)
        tree.column(col, width=width, anchor=anchor)

class CTTCostsViewer:
    def __init__(self, parent):
        self.parent = parent
        self.win = tk.Toplevel(parent)
        self.win.title("CTT Costs")
        self.win.geometry("1200x500")

        self.notebook = ttk.Notebook(self.win)
        self.notebook.pack(expand=True, fill='both')

        self.correo_azul_tab = ttk.Frame(self.notebook, padding=10)
        self.correo_registado_tab = ttk.Frame(self.notebook, padding=10)
        self.ctt_expresso_tab = ttk.Frame(self.notebook, padding=10)
        self.country_mapping_tab = ttk.Frame(self.notebook, padding=10)

        self.notebook.add(self.correo_azul_tab, text="Correio Azul")
        self.notebook.add(self.correo_registado_tab, text="Correio Registado")
        self.notebook.add(self.ctt_expresso_tab, text="CTT Expresso")
        self.notebook.add(self.country_mapping_tab, text="Country Mapping")

        self.available_countries = [
            "Portugal", "Angola", "Saudi Arabia", "Australia", "Brazil", "Canada",
            "Chile", "China", "United Arab Emirates", "France", "Germany", "Ireland",
            "Israel", "Italy", "Japan", "Jordan", "Lithuania", "Morocco", "Mexico",
            "New Zealand", "Poland", "Singapore", "South Africa", "Spain", "Sweden",
            "Switzerland", "Turkey", "United Kingdom", "USA"
        ]

        self.transport_methods = ["Correio Azul", "Correio Registado", "CTT Expresso"]
        self.regions = ["USA Zone", "Rest World Zone", "Europe"]

        self.weight_intervals = [
            "< 0.019",
            "[0.020;0.049]",
            "[0.050;0.099]",
            "[0.100;0.249]",
            "[0.250;0.499]",
            "[0.500;0.999]",
            "[1.000;2.000]"
        ]

        self.azul_data = [
            (self.weight_intervals[0], "4,47", "4,89", "5,23"),
            (self.weight_intervals[1], "4,47", "4,89", "5,23"),
            (self.weight_intervals[2], "4,47", "4,89", "5,23"),
            (self.weight_intervals[3], "5,65", "6", "6,66"),
            (self.weight_intervals[4], "7,55", "10,14", "11,75"),
            (self.weight_intervals[5], "10,65", "17,8", "19,8"),
            (self.weight_intervals[6], "16,75", "28", "30,35")
        ]

        self.registado_data = [
            (self.weight_intervals[0], "5,43", "5,79", "6,11"),
            (self.weight_intervals[1], "5,43", "5,79", "6,11"),
            (self.weight_intervals[2], "5,43", "5,79", "6,11"),
            (self.weight_intervals[3], "6,18", "7,4", "8,47"),
            (self.weight_intervals[4], "8,3", "11,55", "12,13"),
            (self.weight_intervals[5], "11,83", "19,15", "21,2"),
            (self.weight_intervals[6], "17,69", "29,5", "31,7")
        ]

        # New CTT Expresso data
        self.expresso_data = [
            ("Até 1Kg", "3,74"),
            ("Até 5Kg", "4,03"),
            ("Até 10Kg", "4,43"),
            ("Até 15Kg", "4,81"),
            ("Até 20Kg", "5,26"),
            ("Até 25Kg", "5,99"),
            ("Até 30Kg", "6,54"),
            ("Mais de 30Kg", "200")
        ]

        self.azul_tree = self.create_table(self.correo_azul_tab, ["Weight", "Europe", "Rest World", "USA"])
        self.registado_tree = self.create_table(self.correo_registado_tab, ["Weight", "Europe", "Rest World", "USA"])
        self.expresso_tree = self.create_table(self.ctt_expresso_tab, ["Interval", "Price (€)"])
        self.country_mapping_tree = self.create_table(self.country_mapping_tab, ["Country of the Order", "Transport Method", "Region"])

        self.insert_data(self.azul_tree, self.azul_data)
        self.insert_data(self.registado_tree, self.registado_data)
        self.insert_data(self.expresso_tree, self.expresso_data)

        country_mapping = [
            ("United States", "Correio Azul", "USA Zone"),
            ("Canada", "Correio Azul", "Rest World Zone"),
            ("Australia", "Correio Azul", "Rest World Zone"),
            ("New Zealand", "Correio Azul", "Rest World Zone"),
            ("France", "Correio Azul", "Europe")
        ]

        self.insert_data(self.country_mapping_tree, country_mapping)

        self.mapping_controls = ttk.Frame(self.country_mapping_tab)
        self.mapping_controls.pack(fill='x', pady=(0, 10))

        self.add_row_button = ttk.Button(self.mapping_controls, text="Add Row", command=self.add_mapping_row)
        self.add_row_button.pack(side='right', padx=5)

        self.current_tree = None
        self.editing_column = None
        self.editing_item = None
        self.entry_widget = None
        self.combo_widget = None

        self.azul_tree.bind("<Double-1>", self.on_double_click)
        self.registado_tree.bind("<Double-1>", self.on_double_click)
        self.expresso_tree.bind("<Double-1>", self.on_double_click_expresso)
        self.country_mapping_tree.bind("<Double-1>", self.on_mapping_double_click)

        self.save_frame = ttk.Frame(self.win)
        self.save_frame.pack(side='bottom', fill='x', pady=10, padx=10, anchor='e')
        self.save_button = ttk.Button(self.save_frame, text="Save", command=self.save_data)
        self.save_button.pack(side='right')

    def add_mapping_row(self):
        self.country_mapping_tree.insert('', 'end', values=("Select Country", "Correio Azul", "Select Region"))
        last_item = self.country_mapping_tree.get_children()[-1]
        self.country_mapping_tree.selection_set(last_item)
        self.country_mapping_tree.focus(last_item)
        self.on_mapping_double_click(None, force_item=last_item)

    def on_mapping_double_click(self, event, force_item=None):
        if force_item:
            item = force_item
        else:
            if not self.country_mapping_tree.selection():
                return
            item = self.country_mapping_tree.selection()[0]

        column = self.country_mapping_tree.identify_column(event.x) if event else "#1"
        x, y, w, h = self.country_mapping_tree.bbox(item, column)

        values = self.country_mapping_tree.item(item)['values']
        column_id = int(column[1]) - 1

        if self.combo_widget:
            self.combo_widget.destroy()

        if column_id == 0:
            values_list = self.available_countries
        elif column_id == 1:
            values_list = self.transport_methods
        else:
            values_list = self.regions

        self.combo_widget = ttk.Combobox(self.country_mapping_tree, values=values_list)
        self.combo_widget.set(values[column_id])
        self.combo_widget.place(x=x, y=y, width=w, height=h)
        self.combo_widget.focus()

        self.combo_widget.bind('<Return>', lambda e: self.finish_mapping_edit(e, item, column_id))
        self.combo_widget.bind('<FocusOut>', lambda e: self.finish_mapping_edit(e, item, column_id))

        self.current_tree = self.country_mapping_tree
        self.editing_item = item
        self.editing_column = column_id

    def finish_mapping_edit(self, event, item, column):
        if self.combo_widget:
            new_value = self.combo_widget.get()
            current_values = list(self.country_mapping_tree.item(item)['values'])
            current_values[column] = new_value
            self.country_mapping_tree.item(item, values=current_values)
            self.combo_widget.destroy()
            self.combo_widget = None

        self.current_tree = None
        self.editing_item = None
        self.editing_column = None

    def create_table(self, parent_frame, columns):
        tree = ttk.Treeview(parent_frame, columns=columns, show='headings', height=10)

        for col in columns:
            tree.column(col, width=200, anchor=tk.CENTER)
            tree.heading(col, text=col)

        vsb = ttk.Scrollbar(parent_frame, orient="vertical", command=tree.yview)
        tree.configure(yscrollcommand=vsb.set)
        vsb.pack(side='right', fill='y')
        tree.pack(expand=True, fill='both')
        return tree

    def insert_data(self, tree, data):
        for row in data:
            tree.insert('', 'end', values=row)

    def on_double_click(self, event):
        tree = event.widget
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = tree.identify_column(event.x)
        rowid = tree.identify_row(event.y)
        if not rowid:
            return

        col_index = int(column.replace('#', '')) - 1
        if col_index == 0:
            return

        self.current_tree = tree
        self.editing_item = rowid
        self.editing_column = col_index

        current_value = tree.set(rowid, tree["columns"][col_index])
        x, y, w, h = tree.bbox(rowid, column)

        self.entry_widget = ttk.Entry(tree)
        self.entry_widget.insert(0, current_value)
        self.entry_widget.focus()
        self.entry_widget.place(x=x, y=y, width=w, height=h)

        self.entry_widget.bind("<Return>", self.editing_finished)
        self.entry_widget.bind("<FocusOut>", self.editing_finished)

    def on_double_click_expresso(self, event):
        tree = event.widget
        region = tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        column = tree.identify_column(event.x)
        rowid = tree.identify_row(event.y)
        if not rowid:
            return

        col_index = int(column[1]) - 1
        # For CTT Expresso, allow editing price only (col_index == 1)
        if col_index == 0:
            return

        self.current_tree = tree
        self.editing_item = rowid
        self.editing_column = col_index

        current_value = tree.set(rowid, tree["columns"][col_index])
        x, y, w, h = tree.bbox(rowid, column)

        self.entry_widget = ttk.Entry(tree)
        self.entry_widget.insert(0, current_value)
        self.entry_widget.focus()
        self.entry_widget.place(x=x, y=y, width=w, height=h)

        self.entry_widget.bind("<Return>", self.editing_finished)
        self.entry_widget.bind("<FocusOut>", self.editing_finished)

    def editing_finished(self, event):
        if self.entry_widget is None:
            return
        new_value = self.entry_widget.get()
        cols = self.current_tree["columns"]
        self.current_tree.set(self.editing_item, cols[self.editing_column], new_value)
        self.entry_widget.destroy()
        self.entry_widget = None
        self.editing_item = None
        self.editing_column = None
        self.current_tree = None

    def save_data(self):
        print("Correio Azul updated data:")
        for rowid in self.azul_tree.get_children():
            print(self.azul_tree.item(rowid, 'values'))

        print("Correio Registado updated data:")
        for rowid in self.registado_tree.get_children():
            print(self.registado_tree.item(rowid, 'values'))

        print("CTT Expresso updated data:")
        for rowid in self.expresso_tree.get_children():
            print(self.expresso_tree.item(rowid, 'values'))

        print("Country Mapping updated data:")
        for rowid in self.country_mapping_tree.get_children():
            print(self.country_mapping_tree.item(rowid, 'values'))

        messagebox.showinfo("Save", "CTT costs updated successfully!")

class OrderAnalysisGUI:
    def __init__(self, root: tk.Tk):
        self.root = root
        self.root.title("Order Analysis")
        self.root.geometry("1300x900")
        
        self.order_data = []
        self.total_orders = 0
        self.processed_orders = 0
        self.start_time = 0
        self.is_running = False
        self.processed_data = []
        self.group_data = None

        self.style = setup_treeview_style()
        self.main_frame = ttk.Frame(root, padding=30)
        self.main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.create_timer_and_count_widgets()
        self.create_table()
        self.create_progress_section()
        
        style = ttk.Style()
        style.configure('Padded.TButton', padding=(5, 5, 5, 5))

        self.button_frame = ttk.Frame(self.main_frame)
        self.button_frame.grid(row=3, column=0, columnspan=2, sticky=tk.EW, padx=30, pady=30)

        self.ctt_costs_button = ttk.Button(self.button_frame, text="CTT Costs", command=self.ctt_costs_action, style='Padded.TButton')
        self.start_button = ttk.Button(self.button_frame, text="Start Analysis", command=self.start_analysis, style='Padded.TButton')
        self.export_button = ttk.Button(self.button_frame, text="Export Excel", command=self.export_to_excel_main, style='Padded.TButton')
        
        self.ctt_costs_button.pack(side='left', padx=10)
        self.export_button.pack(side='right', padx=10)
        self.start_button.pack(side='right', padx=10)

        self.main_frame.columnconfigure(0, weight=1)
        root.columnconfigure(0, weight=1)
        root.rowconfigure(0, weight=1)

    def create_timer_and_count_widgets(self):
        self.timer_frame = ttk.Frame(self.main_frame, padding=30)
        self.timer_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        self.timer_label = ttk.Label(self.timer_frame, text="Time Elapsed:", font=('Arial', 10, 'bold'))
        self.timer_label.grid(row=0, column=0, padx=30)
        
        self.timer_value = ttk.Label(self.timer_frame, text="00:00", font=('Arial', 12))
        self.timer_value.grid(row=0, column=1, padx=30)
        
        self.order_count_label = ttk.Label(self.timer_frame, text="Orders to Process:", font=('Arial', 10, 'bold'))
        self.order_count_label.grid(row=0, column=2, padx=30)
        
        self.order_count_value = ttk.Label(self.timer_frame, text="0/0", font=('Arial', 12))
        self.order_count_value.grid(row=0, column=3, padx=30)

    def create_table(self):
        self.table_frame = ttk.Frame(self.main_frame, padding=30)
        self.table_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        self.tree = ttk.Treeview(self.table_frame, style='MyTreeview', height=16)
        configure_columns(self.tree)
        
        scrollbar = ttk.Scrollbar(self.table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)
        
        self.tree.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
        self.table_frame.columnconfigure(0, weight=1)

    def create_progress_section(self):
        self.progress_frame = ttk.LabelFrame(self.main_frame, text="Progress", padding=10)
        self.progress_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E))
        
        ttk.Label(self.progress_frame, text="Fetching Orders:").grid(row=0, column=0, sticky=tk.W, padx=10, pady=10)
        self.fetch_progress_var = tk.DoubleVar()
        self.fetch_progress = ttk.Progressbar(
            self.progress_frame, 
            variable=self.fetch_progress_var,
            maximum=100,
            length=300,
            mode='determinate'
        )
        self.fetch_progress.grid(row=0, column=1, padx=10, pady=10)
        self.fetch_progress_label = ttk.Label(self.progress_frame, text="0%")
        self.fetch_progress_label.grid(row=0, column=2, padx=10, pady=10)
        
        ttk.Label(self.progress_frame, text="Processing Data:").grid(row=1, column=0, sticky=tk.W, padx=10, pady=10)
        self.compute_progress_var = tk.DoubleVar()
        self.compute_progress = ttk.Progressbar(
            self.progress_frame,
            variable=self.compute_progress_var,
            maximum=100,
            length=300,
            mode='determinate'
        )
        self.compute_progress.grid(row=1, column=1, padx=10, pady=10)
        self.compute_progress_label = ttk.Label(self.progress_frame, text="0%")
        self.compute_progress_label.grid(row=1, column=2, padx=10, pady=10)

    def update_timer(self):
        if self.is_running:
            elapsed_time = int(time.time() - self.start_time)
            minutes = elapsed_time // 60
            seconds = elapsed_time % 60
            self.timer_value.config(text=f"{minutes:02d}:{seconds:02d}")
            self.root.after(1000, self.update_timer)

    def update_table(self, order_data: List[Dict]):
        try:
            for item in self.tree.get_children():
                self.tree.delete(item)
            
            for i, row in enumerate(order_data):
                values = []
                for col in self.tree["columns"]:
                    val = row.get(col, "")
                    values.append("" if val is None else str(val))
            
                tag = "oddrow" if i % 2 else "evenrow"
                self.tree.insert("", "end", values=values, tags=(tag,))
        
            self.tree.tag_configure("oddrow", background="#F2D5DB")
            self.tree.tag_configure("evenrow", background="#FAEAED")
        
        except Exception as e:
            logger.error(f"Error in update_table: {str(e)}")
            traceback.print_exc()

    def run_analysis(self):
        try:
            orders = fetch_unfulfilled_orders_with_progress(self)
            logger.info("Processing orders now...")
            self.processed_data = self.process_orders_with_progress(orders)
            
            df_group1, df_group2, df_group3 = split_data_by_group(self.processed_data)
            self.group_data = {1: df_group1, 2: df_group2, 3: df_group3}
            
        except Exception as e:
            err_trace = traceback.format_exc()
            logger.error(f"Error during analysis: {str(e)}\n{err_trace}")
            messagebox.showerror("Error", f"An error occurred during analysis:\n{str(e)}")
        finally:
            self.is_running = False
            self.start_button.config(state='normal')

    def process_orders_with_progress(self, orders: List[Dict]) -> List[Dict]:
        processed_data = []
        batch_size = 10
        
        for batch_idx in range(0, len(orders), batch_size):
            batch_orders = orders[batch_idx:batch_idx+batch_size]
            try:
                if batch_orders:
                    logger.debug(f"First order structure: {batch_orders[0]}")
                    print(f"First order structure: {batch_orders[0]}")
                
                orders_data = {'edges': batch_orders}
                batch_processed = process_orders([orders_data])
            except Exception as e:
                err_trace = traceback.format_exc()
                logger.error(f"Error in process_orders_with_progress: {str(e)}\n{err_trace}")
                messagebox.showerror("Error", f"An error occurred in process_orders_with_progress:\n{str(e)}")
                break
            processed_data.extend(batch_processed)
            
            progress = ((batch_idx + len(batch_orders)) / len(orders) * 100) if len(orders) else 0
            self.compute_progress_var.set(progress)
            self.compute_progress_label.config(text=f"{int(progress)}%")
            self.root.update_idletasks()
            
            self.update_table(processed_data[-16:])
        
        return processed_data

    def start_analysis(self):
        self.start_button.config(state='disabled')
        self.start_time = time.time()
        self.is_running = True
        self.update_timer()
        Thread(target=self.run_analysis, daemon=True).start()

    def export_to_excel_main(self):
        try:
            if not self.group_data:
                messagebox.showwarning("Warning", "No data to export. Please run the analysis first.")
                return

            timestamp_str = time.strftime('%Y_%m_%d_%H_%M_%S')
            default_file_name = f"Cosmetyque_Orders_Grouped_{timestamp_str}.xlsx"

            file_path = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                title="Save Excel File",
                initialfile=default_file_name
            )
            
            if not file_path:
                return

            df_group1 = self.group_data[1]
            df_group2 = self.group_data[2]
            df_group3 = self.group_data[3]
            
            save_excel_file(file_path, df_group1, df_group2, df_group3)
            
            messagebox.showinfo("Success", "Data exported successfully!")
            
            try:
                os.startfile(file_path)
            except Exception as e:
                logger.error(f"Error opening Excel file: {str(e)}")
                messagebox.showwarning("Warning", 
                    "File was saved successfully but could not be opened automatically.\n"
                    f"File location: {file_path}")
                
        except Exception as e:
            err_trace = traceback.format_exc()
            logger.error(f"Error exporting to Excel: {str(e)}\n{err_trace}")
            messagebox.showerror("Error", f"An error occurred while exporting to Excel:\n{str(e)}")

    def ctt_costs_action(self):
        CTTCostsViewer(self.root)


def main():
    import socket
    sock = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
    try:
        sock.bind(('localhost', 12345))
        root = tk.Tk()
        app = OrderAnalysisGUI(root)
        root.mainloop()
    except socket.error:
        messagebox.showwarning("Warning", "Application is already running!")
    finally:
        sock.close()

if __name__ == "__main__":
    main()
